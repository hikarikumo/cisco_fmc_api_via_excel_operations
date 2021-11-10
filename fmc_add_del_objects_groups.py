#!/usr/bin/python3
'''
Add/Delete/Modify netobjects, rangeobjects, hostobjects to different domains
Add/Delete/Modify groups with objects
'''
from get_func import get_all_domains_data, get_all_objects_with_domains, get_all_objects_for_domain, get_all_groups_info, get_all_devices, get_domain_uuid, get_object_data, check_if_object_already_exist, get_all_detailed_groups_for_domain, get_all_objects_for_domain_no_check_ids
from put_func import post_network_objects, post_range_objects, post_host_objects, post_groups_objects, del_groups, put_networkgroups, del_objects, put_object, post_url_objects, put_urlgroups, post_urlgroups_objects, del_urlgroups
from fmc_create_xlsx_obj_groups import create_xlsx_and_sheets, write_group_objects_to_xlsx, write_hosts_network_objects_to_xlsx, write_urlgrps_to_xlsx
from fmc_excel_diff_input_output import create_diff_excel_file
import pandas as pd
import json
import cfg
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
import re
import copy
import os
from datetime import datetime
import time



logging.basicConfig(format='%(threadName)s %(name)s %(levelname)s: %(message)s',level=logging.INFO)


def check_parent_group_non_override(obj, domain_name):
    all_detailed_networkgroups = cfg.all_detailed_networkgroups
    parent_groups_dict = dict()
    for networkgroup, networkgroup_data in all_detailed_networkgroups[domain_name].items():
        if not networkgroup_data['overridable']:
            try:
                for index, object in enumerate(networkgroup_data['objects']):
                    if object['name'] == obj:
                        parent_groups_dict.update({networkgroup: index})                       
            except KeyError as error:
                logging.info(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} networkgroup_data objects of {networkgroup_data} do not exist')
                errors_filename = 'outputs/errors.txt'
                with open(errors_filename, "a") as f:
                    f.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} networkgroup_data objects of {networkgroup_data} do not exist\n Error: {error}\n')
    # logging.info(f'networkgroups = {parent_groups_dict}')
    return parent_groups_dict


def check_parent_urlgroup(obj, domain_name):
    parent_groups_dict = dict()
    for urlgroup, urlgroup_data in cfg.all_obj_domain.get(domain_name).get('urlgroups').items():
        if not urlgroup_data['overridable']:
            try:
                for index, object in enumerate(urlgroup_data['objects']):
                    if object['name'] == obj:
                        parent_groups_dict.update({urlgroup: index})
            except KeyError as error:
                logging.info(
                    f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} urlgroup_data objects of {urlgroup_data} do not exist')
                errors_filename = 'outputs/errors.txt'
                with open(errors_filename, "a") as f:
                    f.write(
                        f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} urlgroup_data objects of {urlgroup_data} do not exist\n Error: {error}\n')
    # logging.info(f'urlgroups = {parent_groups_dict}')
    return parent_groups_dict


def create_del_json_non_override(parent_group_name, obj_del_index, obj, domain_name):
    all_detailed_networkgroups = cfg.all_detailed_networkgroups
    
    stripped_json_for_put = dict()
    old_parent_group_json = dict()
    new_parent_group_json = dict()
    old_parent_group_json = all_detailed_networkgroups[domain_name][parent_group_name]
    del old_parent_group_json['objects'][obj_del_index]
    new_parent_group_json = old_parent_group_json
    
    stripped_json_for_put.update({'id': new_parent_group_json['id']})
    stripped_json_for_put.update({'overridable': new_parent_group_json['overridable']})
    stripped_json_for_put.update({'description': new_parent_group_json['description']})
    stripped_json_for_put.update({'objects': new_parent_group_json['objects']})
    stripped_json_for_put.update({'name': new_parent_group_json['name']})
    stripped_json_for_put.update({'type': new_parent_group_json['type']})

    return stripped_json_for_put


def create_del_json_urlgrp_non_override(parent_group_name, obj_del_index, obj, domain_name):
        
    stripped_json_for_put = dict()
    old_parent_group_json = dict()
    new_parent_group_json = dict()
    old_parent_group_json = cfg.all_obj_domain[domain_name]['urlgroups'][parent_group_name]
    del old_parent_group_json['objects'][obj_del_index]
    new_parent_group_json = old_parent_group_json
    
    stripped_json_for_put.update({'id': new_parent_group_json['id']})
    stripped_json_for_put.update({'overridable': new_parent_group_json['overridable']})
    stripped_json_for_put.update({'description': new_parent_group_json['description']})
    stripped_json_for_put.update({'objects': new_parent_group_json['objects']})
    stripped_json_for_put.update({'name': new_parent_group_json['name']})
    stripped_json_for_put.update({'type': new_parent_group_json['type']})

    return stripped_json_for_put


def check_if_object_value_changed(object_name, obj_value, domain_name):
    all_obj_domain = cfg.all_obj_domain
    system_objects = cfg.system_objects

    if object_name in system_objects:
        return False

    if all_obj_domain:
        # object_type = check_object_type(object_name)
        for object_type in cfg.object_types:
            if all_obj_domain[domain_name].get(object_type).get(object_name):
                if object_type == 'urls':
                    object_value_in_fmc = all_obj_domain[domain_name][object_type][object_name].get('url')
                else:
                    object_value_in_fmc = all_obj_domain[domain_name][object_type][object_name].get('value')
                if object_value_in_fmc != obj_value:
                    return True
    

def del_group_from_parent_group(domain_name, ws):
    """
    del_group_from_parent_group delete group from parent group.
    1st: check whether the group (to be deleted) is within any group, if no - just delete it
    2nd: if in group, get group elements (method: all_detailed_networkgroups = get_all_groups_info())
    3rd: create json of the parent group without the group (to be deleted)
    4th: PUT via API the created json via request to the parent group ID
    5th: in case the parent group consists from only one element(which is to be deleted), then delete the whole parent group
    
    :param domain_name: domain name
    :type domain_name: str()
    :param ws: sheet which has groups and objects listed
    :type ws: openpyxl class object
    """
    # OBJECT_HOST_NAME_START = cfg.OBJECT_HOST_NAME_START
    # OBJECT_RANGE_NAME_START = cfg.OBJECT_RANGE_NAME_START
    # OBJECT_SUBNET_NAME_START = cfg.OBJECT_SUBNET_NAME_START
    # OBJECT_GROUP_NAME_START = cfg.OBJECT_GROUP_NAME_START
    
    
    try:
        uuid = get_domain_uuid(domain_name)['uuid']
    except TypeError:
        errors_filename = 'outputs/errors.txt'
        with open(errors_filename, "a") as f:
            f.write(f'Domain {domain_name} do not exist\n')

    ''' get size of xlsx sheet objects'''
    max_row = ws.max_row
    max_column = ws.max_column
    
    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1      
    
    
    add_modify_non_override_payload = dict()
    add_delete_payload = dict()
   
    for i in range(2, max_row+1):
        group_name = ws.cell(row=i, column=1).value
        obj = ws.cell(row=i, column=2).value
        action = ws.cell(row=i, column=3).value
        
        if group_name:
            group_name = group_name.strip()
        if obj:
            obj = obj.strip()
        if action:
            action = action.strip()
        
        if action:
            del_group_in_non_override_group = [
                obj, 
                action == 'delete']
            
            if all(del_group_in_non_override_group):
                # if obj.startswith(OBJECT_GROUP_NAME_START):
                parent_group = check_parent_group_non_override(
                    obj, domain_name)
                if parent_group:
                    for parent_group_name, obj_del_index in parent_group.items():
                        if parent_group_name == group_name:
                            new_parent_group_json = create_del_json_non_override(
                                parent_group_name, obj_del_index, obj, domain_name)
                            if len(new_parent_group_json['objects']) > 0:
                                add_modify_non_override_payload.update({new_parent_group_json['name']: new_parent_group_json})
                            else:
                                add_delete_payload.update({new_parent_group_json['name']: new_parent_group_json})

    if add_modify_non_override_payload:
        for group_name, group_data in add_modify_non_override_payload.items():
            if len(group_data['objects']) > 0:
                try:
                    put_networkgroups(group_data, domain_name)
                except KeyError:
                    logging.info(
                        f'domain = {domain_name} do not has Group objects to modify')
    if add_delete_payload:
        for group_name, group_data in add_delete_payload.items():
            try:
                del_groups(group_data, domain_name)
            except KeyError:
                logging.info(
                    f'domain = {domain_name} do not has Group objects to modify')
    # logging.info(f"\n{50*'#'}"
    #              f'\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished checking DEL request to groups in {domain_name}\n'
    #              f"{50*'#'}")
    if any([add_modify_non_override_payload, add_delete_payload]):
        ''' update all_objects cause '''
        update_all_objects(domain_name)
        update_all_networkgroups(domain_name)


def do_ask_user_input(user_prompt):
    """
    Default action - No
    Return y or n
    """
    input_y_or_n = 'user input'
    while (input_y_or_n != 'y' and input_y_or_n != 'n'):
        input_y_or_n = input(user_prompt).lower()
        if not input_y_or_n:
            # Enter
            input_y_or_n = 'n'
    return input_y_or_n


def remove_space_obj(obj):
    obj_space_match = re.search(
        r'(?P<begin>.+?(?=\s+))'
        r'\s+'
        r'(?P<end>.+)', obj)
    if obj_space_match:
        obj = f'{obj_space_match.group("begin")}{obj_space_match.group("end")}'
    return obj


def get_group_data_from_detailed_networgroups(group_names, domain_name):
    all_detailed_networkgroups = cfg.all_detailed_networkgroups
    
    tmp_dict = dict()
    for group in group_names:
        group_objects = all_detailed_networkgroups.get(domain_name).get(group).get('objects')
        tmp_list = [item.get('name') for item in group_objects]
        tmp_dict.setdefault(group, sorted(tmp_list))
    return tmp_dict


def compare_group_data_to_avoid_uneeded_puts(ws, domain_name):
    """
    Open excel sheet data as pandas DataFrame. Remove first row out of the DataFrame as data row. Use first row as columns name for the DataFrame.
    """
    df = pd.DataFrame(ws.values)
    new_header = df.iloc[0]
    df = df[1:]
    df = df.rename(columns=new_header)

    """ temp list to include all data exported out of the DataFrame """
    list_sorted_groups_False_Add = list()
    dict_sorted_groups_for_comparison = dict()
    df_groups_sort = df.groupby(['object_group_name', 'action'])

    for group, group_data1 in df_groups_sort:
        if (group[1] == 'add'):
            temp_group = df_groups_sort.get_group(group)
            temp = temp_group.to_dict('records')
            list_sorted_groups_False_Add.append(temp)
            tmp_list = [item.get('object') for item in temp]
            dict_sorted_groups_for_comparison.update({group[0]: sorted(tmp_list)})
    
    group_names = list(dict_sorted_groups_for_comparison)
    
    dict_group_data_from_detailed_networgroups = get_group_data_from_detailed_networgroups(group_names, domain_name)
    different_groups = list()
    for group, group_data in dict_sorted_groups_for_comparison.items():
        if group_data == dict_group_data_from_detailed_networgroups.get(group):
            continue
        else:
            different_groups.append(group)
    return different_groups    
        

def change_hosts_nets_in_parent_group(domain_name, ws):
    """
    del_hosts_nets_from_parent_group delete group from parent group.
    1st: check whether the group (to be deleted) is within any group, if no - just delete it
    2nd: if in group, get group elements (method: all_detailed_networkgroups = get_all_groups_info())
    3rd: create json of the parent group without the group (to be deleted)
    4th: PUT via API the created json via request to the parent group ID
    5th: in case the parent group consists from only one element(which is to be deleted), then delete the whole parent group
    
    :param domain_name: domain name
    :type domain_name: str()
    :param ws: sheet which has groups and objects listed
    :type ws: openpyxl class object
    """
    # OBJECT_HOST_NAME_START = cfg.OBJECT_HOST_NAME_START
    # OBJECT_RANGE_NAME_START = cfg.OBJECT_RANGE_NAME_START
    # OBJECT_SUBNET_NAME_START = cfg.OBJECT_SUBNET_NAME_START
    # OBJECT_GROUP_NAME_START = cfg.OBJECT_GROUP_NAME_START
    
    try:
        uuid = get_domain_uuid(domain_name)['uuid']
    except TypeError as error:
        logging.info(f'Domain {domain_name} do not exist either on FMC, either in Excel sheet')
        errors_filename = 'outputs/errors.txt'
        with open(errors_filename, "a") as f:
            f.write(f'Domain {domain_name} do not exist either on FMC, either in Excel sheet\n Error: {error}\n')

    ''' get size of xlsx sheet objects'''
    max_row = ws.max_row
    max_column = ws.max_column

    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1

    add_modify_non_override_payload = dict()
    add_delete_payload = dict()

    # logging.info(f"\n{50*'#'}"
    #              f'\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to check Modify/DEL request to exclude objects from groups in {domain_name}\n'
    #              f"{50*'#'}")
    for i in range(2, max_row+1):
        group_name = ws.cell(row=i, column=1).value
        obj = ws.cell(row=i, column=2).value
        action = ws.cell(row=i, column=3).value

        if group_name:
            group_name = group_name.strip()
        if obj:
            obj = obj.strip()
        if action:
            action = action.strip()
        if action:
            del_obj_in_non_override_group = [
                obj,
                action == 'delete']
            if all(del_obj_in_non_override_group):
                # if any([obj.startswith(OBJECT_HOST_NAME_START), obj.startswith(OBJECT_SUBNET_NAME_START), obj.startswith(OBJECT_RANGE_NAME_START)]):
                parent_group = check_parent_group_non_override(
                    obj, domain_name)
                if parent_group:
                    for parent_group_name, obj_del_index in parent_group.items():
                        if parent_group_name == group_name:
                            new_parent_group_json = create_del_json_non_override(
                                parent_group_name, obj_del_index, obj, domain_name)
                            if len(new_parent_group_json['objects']) > 0:
                                add_modify_non_override_payload.update(
                                    {new_parent_group_json['name']: new_parent_group_json})
                            else:
                                add_delete_payload.update(
                                    {new_parent_group_json['name']: new_parent_group_json})
    if add_modify_non_override_payload:
        for group_name, group_data in add_modify_non_override_payload.items():
            if len(group_data['objects']) > 0:
                try:
                    put_networkgroups(group_data, domain_name)
                except KeyError:
                    logging.info(
                        f'domain = {domain_name} do not has Group objects to modify')
    if add_delete_payload:
        for group_name, group_data in add_delete_payload.items():
            try:
                del_groups(group_data, domain_name)
            except KeyError:
                logging.info(
                    f'domain = {domain_name} do not has Group objects to modify')
    # logging.info(f"\n{50*'#'}"
    #              f'\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished checking DEL request to groups in {domain_name}\n'
    #              f"{50*'#'}")
    if any([add_modify_non_override_payload, add_delete_payload]):
        ''' update all_objects cause new hosts have been just added '''
        update_all_objects(domain_name)
        update_all_networkgroups(domain_name)


def change_urls_in_group(domain_name, ws):
    """
    change_urls_in_group delete group from parent group.
    1st: check whether the group (to be deleted) is within any group, if no - just delete it
    2nd: if in group, get group elements (method: all_detailed_networkgroups = get_all_groups_info())
    3rd: create json of the parent group without the group (to be deleted)
    4th: PUT via API the created json via request to the parent group ID
    5th: in case the parent group consists from only one element(which is to be deleted), then delete the whole parent group
    
    :param domain_name: domain name
    :type domain_name: str()
    :param ws: sheet which has groups and objects listed
    :type ws: openpyxl class object
    """
    try:
        uuid = get_domain_uuid(domain_name)['uuid']
    except TypeError as error:
        logging.info(
            f'Domain {domain_name} do not exist either on FMC, either in Excel sheet')
        errors_filename = 'outputs/errors.txt'
        with open(errors_filename, "a") as f:
            f.write(
                f'Domain {domain_name} do not exist either on FMC, either in Excel sheet\n Error: {error}\n')

    ''' get size of xlsx sheet objects'''
    max_row = ws.max_row
    max_column = ws.max_column

    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1

    add_modify_non_override_payload = dict()
    add_delete_payload = dict()

    # logging.info(f"\n{50*'#'}"
    #              f'\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to check Modify/DEL request to exclude objects from groups in {domain_name}\n'
    #              f"{50*'#'}")
    for i in range(2, max_row+1):
        group_name = ws.cell(row=i, column=1).value
        obj = ws.cell(row=i, column=2).value
        action = ws.cell(row=i, column=3).value

        if group_name:
            group_name = group_name.strip()
        if obj:
            obj = obj.strip()
        if action:
            action = action.strip()
        if action:
            del_obj_in_non_override_group = [
                obj,
                action == 'delete']
            if all(del_obj_in_non_override_group):
                # if any([obj.startswith(OBJECT_HOST_NAME_START), obj.startswith(OBJECT_SUBNET_NAME_START), obj.startswith(OBJECT_RANGE_NAME_START)]):
                parent_group = check_parent_urlgroup(
                    obj, domain_name)
                if parent_group:
                    for parent_group_name, obj_del_index in parent_group.items():
                        if parent_group_name == group_name:
                            new_parent_group_json = create_del_json_urlgrp_non_override(
                                parent_group_name, obj_del_index, obj, domain_name)
                            if len(new_parent_group_json['objects']) > 0:
                                add_modify_non_override_payload.update(
                                    {new_parent_group_json['name']: new_parent_group_json})
                            else:
                                add_delete_payload.update(
                                    {new_parent_group_json['name']: new_parent_group_json})
    if add_modify_non_override_payload:
        for group_name, group_data in add_modify_non_override_payload.items():
            if len(group_data['objects']) > 0:
                try:
                    put_urlgroups(group_data, domain_name)
                except KeyError:
                    logging.info(
                        f'domain = {domain_name} do not has Group objects to modify')
    if add_delete_payload:
        for group_name, group_data in add_delete_payload.items():
            try:
                del_urlgroups(group_data, domain_name)
            except KeyError:
                logging.info(
                    f'domain = {domain_name} do not has Group objects to modify')
    if any([add_modify_non_override_payload]):
        ''' update all_objects cause new objects were added '''
        update_all_objects(domain_name)

        

def del_hosts_nets_for_obj_sheet(domain_name, ws):
    """
    del_hosts_nets_from_parent_group delete group from parent group.
    1st: check whether the group (to be deleted) is within any group, if no - just delete it
    2nd: if in group, get group elements (method: all_detailed_networkgroups = get_all_groups_info())
    3rd: create json of the parent group without the group (to be deleted)
    4th: PUT via API the created json via request to the parent group ID
    5th: in case the parent group consists from only one element(which is to be deleted), then delete the whole parent group
    
    :param domain_name: domain name
    :type domain_name: str()
    :param ws: sheet which has groups and objects listed
    :type ws: openpyxl class object
    """
    # OBJECT_HOST_NAME_START = cfg.OBJECT_HOST_NAME_START
    # OBJECT_RANGE_NAME_START = cfg.OBJECT_RANGE_NAME_START
    # OBJECT_SUBNET_NAME_START = cfg.OBJECT_SUBNET_NAME_START
    # OBJECT_GROUP_NAME_START = cfg.OBJECT_GROUP_NAME_START
    
    try:
        uuid = get_domain_uuid(domain_name)['uuid']
    except TypeError as error:
        logging.info(f'Domain {domain_name} do not exist either on FMC, either in Excel sheet')
        errors_filename = 'outputs/errors.txt'
        with open(errors_filename, "a") as f:
            f.write(f'Domain {domain_name} do not exist either on FMC, either in Excel sheet\n Error: {error}\n')

    ''' get size of xlsx sheet objects'''
    max_row = ws.max_row
    max_column = ws.max_column

    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1
            
    add_modify_non_override_payload = dict()
    add_delete_payload = dict()
    add_delete_obj_payload = dict()

    """
    Open excel sheet data as pandas DataFrame. Remove first row out of the DataFrame as data row. Use first row as columns name for the DataFrame.
    """    
    df = pd.DataFrame(ws.values)
    new_header = df.iloc[0]
    df = df[1:]
    df = df.rename(columns=new_header)
    
    """ temp list to include all data exported out of the DataFrame """
    list_sorted_groups_Del = list()
    
    df_groups_sort = df.groupby(['object_name', 'action'])
    
    for group, group_data1 in df_groups_sort:
        if (group[1] == 'delete'):
            temp_group = df_groups_sort.get_group(group)
            temp = temp_group.to_dict('records')
            list_sorted_groups_Del.append(temp)
           
    for groups_desc in list_sorted_groups_Del:
        check_object_name = groups_desc[0].get('object_name')
        # if any([check_object_name.startswith(OBJECT_HOST_NAME_START), check_object_name.startswith(OBJECT_SUBNET_NAME_START), check_object_name.startswith(OBJECT_RANGE_NAME_START)]):
        parent_group = check_parent_group_non_override(check_object_name, domain_name)
        if parent_group:
            logging.info(f'Warning! Object {check_object_name} is binded with group {parent_group}. Please remove it from the group membership first!')
            errors_filename = 'outputs/errors.txt'
            with open(errors_filename, "a") as f:
                f.write(f'Warning! Object {check_object_name} is binded with group {parent_group}. Please remove it from the group membership first!')
            continue
        else:
            try:
                if check_if_object_already_exist(check_object_name, domain_name):
                    object_name, object_data = get_object_data(check_object_name, domain_name)
                    add_delete_obj_payload.update({object_name: object_data})
            except KeyError as error:
                logging.info(f'Object {check_object_name} do not exist')
                errors_filename = 'outputs/errors.txt'
                with open(errors_filename, "a") as f:
                    f.write(f'Object {check_object_name} do not exist\n Error: {error}\n')

    if add_modify_non_override_payload:
        for group_name, group_data in add_modify_non_override_payload.items():
            if len(group_data['objects']) > 0:
                try:
                    put_networkgroups(group_data, domain_name)
                except KeyError:
                    logging.info(
                        f'domain = {domain_name} do not has Group objects to modify')

    if add_delete_payload:
        for group_name, group_data in add_delete_payload.items():
            try:
                del_groups(group_data, domain_name)
            except KeyError:
                logging.info(
                    f'domain = {domain_name} do not has Group objects to modify')
    if add_delete_obj_payload:
        for obj_name, obj_data in add_delete_obj_payload.items():
            try:
                del_objects(obj_data, domain_name)
            except KeyError:
                logging.info(
                    f'domain = {domain_name} do not has Group objects to modify')
    # logging.info(f"\n{50*'#'}"
    #              f'\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished checking DEL request hosts, networks in {domain_name}\n'
    #              f"{50*'#'}")
    if any([add_modify_non_override_payload, add_delete_payload]):
        ''' update all_objects cause new hosts have been just added '''
        update_all_objects(domain_name)
        update_all_networkgroups(domain_name)


def match_space(object_name):
    space_match = re.search(
        r'\s+', object_name
    )
    if space_match:
        return True
    else:
        return False


def object_change_function(domain_name, ws):
    """
    object_change_function function to add, delete objects

    :param domain_name: domain name of the domain to change object
    :type domain_name: str
    :param ws: Excel sheet where objects are provided
    :type ws: openpyxl class
    """
    all_obj_domain = cfg.all_obj_domain
    
    domains_add_hostobj = dict()
    domains_add_netobj = dict()
    domains_add_rangeobj = dict()
    domains_add_urlobj = dict()
    domains_modify_hostobj = dict()
    domains_modify_netobj = dict()
    domains_modify_rangeobj = dict()
    domains_modify_urlobj = dict()
  
    max_row = ws.max_row
    max_column = ws.max_column

    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1

    add_hostobject_payload = list()
    add_netobject_payload = list()
    add_rangeobject_payload = list()
    add_urlobject_payload = list()
    modify_hostobject_payload = list()
    modify_netobject_payload = list()
    modify_rangeobject_payload = list()
    modify_urlobject_payload = list()
    
    # logging.info(f"\n{50*'#'}"
    #              f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to check addings hosts, ranges, networks to {domain_name}\n'
    #              f"{50*'#'}")
    
    for i in range(2, max_row+1):
        object_name = ws.cell(row=i, column=1).value
        if object_name:
            object_name = object_name.strip()
            if match_space(object_name):
                print(f'SPACE is present in object_name name {object_name}')
                remove_space = do_ask_user_input('Remove SPACE from the object_name [y/N] ') == 'y'
                if remove_space:
                    while (match_space(object_name)):
                        object_name = remove_space_obj(object_name)

        obj = ws.cell(row=i, column=2).value
        object_action = ws.cell(row=i, column=3).value
        object_type = ws.cell(row=i, column=4).value
        if object_action:
            object_action = object_action.lower()
        if object_type:
            object_type = object_type.lower()
        if object_name:
            object_name = object_name.strip()
        # object_type = check_object_type(object_name)
        if obj:
            obj = obj.strip()
            host_add_non_exist = all(
                [
                #  check_object_type(object_name) == 'host',
                object_type == 'host',
                not check_if_object_already_exist(object_name, domain_name),
                 object_action == 'add'])            
            network_add_non_exist = all(
                [
                # check_object_type(object_name) == 'network',
                object_type == 'network',
                not check_if_object_already_exist(object_name, domain_name),
                 object_action == 'add'])
            range_add_non_exist = all(
                [
                # check_object_type(object_name) == 'range',
                object_type == 'range',
                not check_if_object_already_exist(object_name, domain_name),
                 object_action == 'add'])
            url_add_non_exist = all(
                [
                # check_object_type(object_name) == 'range',
                object_type == 'url',
                not check_if_object_already_exist(object_name, domain_name),
                 object_action == 'add'])
            host_add_exist = all(
                [
                #  check_object_type(object_name) == 'host',
                object_type == 'host',
                check_if_object_already_exist(object_name, domain_name),
                 object_action == 'modify'])            
            network_add_exist = all(
                [
                    # check_object_type(object_name) == 'network',
                 object_type == 'network',
                 check_if_object_already_exist(object_name, domain_name),
                 object_action == 'modify'])
            range_add_exist = all(
                [
                # check_object_type(object_name) == 'range',
                object_type == 'range',
                check_if_object_already_exist(object_name, domain_name),
                object_action == 'modify'])
            url_add_exist = all(
                [
                # check_object_type(object_name) == 'range',
                object_type == 'url',
                check_if_object_already_exist(object_name, domain_name),
                object_action == 'modify'])

            if host_add_non_exist:
                hostObject = f'{{"name": "{object_name}","value": "{obj}","type": "{object_type}"}}'
                add_hostobject_payload.append(json.loads(hostObject))
                
            elif network_add_non_exist:
                netObject = f'{{"name": "{object_name}","value": "{obj}","type": "{object_type}"}}'
                add_netobject_payload.append(json.loads(netObject))

            elif range_add_non_exist:
                rangeObject = f'{{"name": "{object_name}","value": "{obj}","type": "{object_type}"}}'
                add_rangeobject_payload.append(json.loads(rangeObject))

            elif url_add_non_exist:
                urlObject = f'{{"name": "{object_name}","url": "{obj}","type": "{object_type}"}}'
                add_urlobject_payload.append(json.loads(urlObject))

            elif host_add_exist:
                if check_if_object_value_changed(object_name, obj, domain_name):
                    hostObject = f'{{"name": "{object_name}","value": "{obj}","type": "{object_type}","id": "{all_obj_domain[domain_name][f"{object_type}s"].get(object_name).get("id")}"}}'
                    modify_hostobject_payload.append(json.loads(hostObject))
            
            elif network_add_exist:
                if check_if_object_value_changed(object_name, obj, domain_name):
                    netObject = f'{{"name": "{object_name}","value": "{obj}","type": "{object_type}","id": "{all_obj_domain[domain_name][f"{object_type}s"].get(object_name).get("id")}"}}'
                    modify_netobject_payload.append(json.loads(netObject))
            
            elif range_add_exist:
                if check_if_object_value_changed(object_name, obj, domain_name):
                    rangeObject = f'{{"name": "{object_name}","value": "{obj}","type": "{object_type}","id": "{all_obj_domain[domain_name][f"{object_type}s"].get(object_name).get("id")}"}}'
                    modify_rangeobject_payload.append(json.loads(rangeObject))
            
            elif url_add_exist:
                if check_if_object_value_changed(object_name, obj, domain_name):
                    urlObject = f'{{"name": "{object_name}","url": "{obj}","type": "{object_type}","id": "{all_obj_domain[domain_name][f"{object_type}s"].get(object_name).get("id")}"}}'
                    modify_urlobject_payload.append(json.loads(urlObject))

    if add_hostobject_payload:
        domains_add_hostobj.update({domain_name: {'hostobject': add_hostobject_payload}})
    if add_netobject_payload:
        domains_add_netobj.update({domain_name: {'netobject': add_netobject_payload}})
    if add_rangeobject_payload:
        domains_add_rangeobj.update({domain_name: {'rangeobject': add_rangeobject_payload}})
    if add_urlobject_payload:
        domains_add_urlobj.update({domain_name: {'urlobject': add_urlobject_payload}})        

    if modify_hostobject_payload:
        domains_modify_hostobj.update({domain_name: {'hostobject': modify_hostobject_payload}})
    if modify_netobject_payload:
        domains_modify_netobj.update({domain_name: {'netobject': modify_netobject_payload}})
    if modify_rangeobject_payload:
        domains_modify_rangeobj.update({domain_name: {'rangeobject': modify_rangeobject_payload}})
    if modify_urlobject_payload:
        domains_modify_urlobj.update({domain_name: {'urlobject': modify_urlobject_payload}})
        
    ''' now to POST our list of network objects '''
    for domain, domain_data in domains_add_netobj.items():
        try:
            if domain_data:
                if len(domain_data['netobject']) < 1000:
                    logging.info(f'Starting to add {len(domain_data["netobject"])} network objects')
                    post_network_objects(domain_data['netobject'], get_domain_uuid(domain)['uuid'])
                elif len(domain_data['netobject']) >= 1000:
                    max_chunks = int(len(domain_data['netobject'])/1000 + 1)
                    for item in range(max_chunks):
                        globals()['net_chunk_list%s' % item] = list()
                    chunk_size_counter = 0
                    chunk_counter = 0
                    for index, item in enumerate(domain_data['netobject']):
                        if index == 1000 + chunk_size_counter:
                            chunk_size_counter += 1000
                            chunk_counter += 1
                        if index < 1000 + chunk_size_counter:
                            globals()['net_chunk_list%s' % chunk_counter].append(item)
                    for chunk in range(max_chunks):
                        logging.info(f'Starting to add {len(globals()["net_chunk_list%s" % chunk])} network objects')
                        post_network_objects(globals()['net_chunk_list%s' % chunk], get_domain_uuid(domain)['uuid'])                

        except KeyError:
            logging.info(
                f'domain = {domain} do not has Network objects to add')
            pass
    ''' now to POST our list of range objects '''
    for domain, domain_data in domains_add_rangeobj.items():
        try:
            if domain_data:
                if len(domain_data['rangeobject']) < 1000:
                    logging.info(f'Starting to add {len(domain_data["rangeobject"])} range objects')
                    post_range_objects(domain_data['rangeobject'], get_domain_uuid(domain)['uuid'])
                elif len(domain_data['rangeobject']) >= 1000:
                    max_chunks = int(len(domain_data['rangeobject'])/1000 + 1)
                    for item in range(max_chunks):
                        globals()['range_chunk_list%s' % item] = list()
                    chunk_size_counter = 0
                    chunk_counter = 0
                    for index, item in enumerate(domain_data['rangeobject']):
                        if index == 1000 + chunk_size_counter:
                            chunk_size_counter += 1000
                            chunk_counter += 1
                        if index < 1000 + chunk_size_counter:
                            globals()['range_chunk_list%s' % chunk_counter].append(item)
                    for chunk in range(max_chunks):
                        logging.info(f'Starting to add {len(globals()["range_chunk_list%s" % chunk])} range objects')
                        post_range_objects(globals()['range_chunk_list%s' % chunk], get_domain_uuid(domain)['uuid'])
        except KeyError:
            logging.info(f'domain = {domain} do not has Range objects to add')
            pass
    ''' now to POST our list of host objects '''
    for domain, domain_data in domains_add_hostobj.items():
        try:
            if domain_data:
                if len(domain_data['hostobject']) < 1000:
                    logging.info(f'Starting to add {len(domain_data["hostobject"])} host objects')
                    post_host_objects(domain_data['hostobject'], get_domain_uuid(domain)['uuid'])
                elif len(domain_data['hostobject']) >= 1000:
                    max_chunks = int(len(domain_data['hostobject'])/1000 + 1)
                    for item in range(max_chunks):
                        globals()['host_chunk_list%s' % item] = list()
                    chunk_size_counter = 0
                    chunk_counter = 0
                    for index, item in enumerate(domain_data['hostobject']):
                        if index == 1000 + chunk_size_counter:
                            chunk_size_counter += 1000
                            chunk_counter += 1
                        if index < 1000 + chunk_size_counter:
                            globals()['host_chunk_list%s' % chunk_counter].append(item)
                    for chunk in range(max_chunks):
                        logging.info(f'Starting to add {len(globals()["host_chunk_list%s" % chunk])} host objects')
                        post_host_objects(globals()['host_chunk_list%s' % chunk], get_domain_uuid(domain)['uuid'])
        except KeyError:
            logging.info(f'domain = {domain} do not has Host objects to add')
    ''' now to POST our list of url objects '''
    for domain, domain_data in domains_add_urlobj.items():
        try:
            if domain_data:
                if len(domain_data['urlobject']) < 1000:
                    logging.info(
                        f'Starting to add {len(domain_data["urlobject"])} url objects')
                    post_url_objects(
                        domain_data['urlobject'], get_domain_uuid(domain)['uuid'])
                elif len(domain_data['urlobject']) >= 1000:
                    max_chunks = int(len(domain_data['urlobject'])/1000 + 1)
                    for item in range(max_chunks):
                        globals()['url_chunk_list%s' % item] = list()
                    chunk_size_counter = 0
                    chunk_counter = 0
                    for index, item in enumerate(domain_data['urlobject']):
                        if index == 1000 + chunk_size_counter:
                            chunk_size_counter += 1000
                            chunk_counter += 1
                        if index < 1000 + chunk_size_counter:
                            globals()['url_chunk_list%s' %
                                      chunk_counter].append(item)
                    for chunk in range(max_chunks):
                        logging.info(
                            f'Starting to add {len(globals()["url_chunk_list%s" % chunk])} url objects')
                        post_url_objects(
                            globals()['url_chunk_list%s' % chunk], get_domain_uuid(domain)['uuid'])

        except KeyError:
            logging.info(
                f'domain = {domain} do not has Network objects to add')
            pass
    ''' now to PUT our list of network objects '''
    for domain, domain_data in domains_modify_netobj.items():
        try:
            if domain_data:
                for item in domain_data.get('netobject'):
                    put_object(item,get_domain_uuid(domain)['uuid'])
        except KeyError:
            logging.info(
                f'domain = {domain} do not has Network objects to modify')
            pass
    for domain, domain_data in domains_modify_hostobj.items():
        try:
            if domain_data:
                for item in domain_data.get('hostobject'):
                    put_object(item,get_domain_uuid(domain)['uuid'])
        except KeyError:
            logging.info(
                f'domain = {domain} do not has Host objects to modify')
            pass
    for domain, domain_data in domains_modify_rangeobj.items():
        try:
            if domain_data:
                for item in domain_data.get('rangeobject'):
                    put_object(item,get_domain_uuid(domain)['uuid'])
        except KeyError:
            logging.info(
                f'domain = {domain} do not has Range objects to modify')
            pass
    for domain, domain_data in domains_modify_urlobj.items():
        try:
            if domain_data:
                for item in domain_data.get('urlobject'):
                    put_object(item, get_domain_uuid(domain)['uuid'])
        except KeyError:
            logging.info(
                f'domain = {domain} do not has Url objects to modify')
            pass
        
    if any([add_netobject_payload, add_rangeobject_payload, add_hostobject_payload, add_urlobject_payload]):
        ''' update all_objects cause new hosts have been just added '''
        update_all_objects(domain_name)
        update_all_networkgroups(domain_name)
   

def create_json_put_obj(group_data, group_object_data, override_object=False):
    group_data_for_json = {}
    if override_object:
        pass
    else:
        new_objects = list()
        for item in group_object_data:
            new_objects.append(item)
        for element in cfg.all_detailed_networkgroups[group_data['domain_name']][group_data['name']]['objects']:
            new_objects.append(element)
        group_data_for_json.update({"objects": new_objects})
        group_data_for_json.update({"type": "NetworkGroup"})
        group_data_for_json.update({"name": group_data['name']})
        object_name, object_data = get_object_data(group_data['name'], group_data['domain_name'])
        group_data_for_json.update({"id": object_data['id']})
    return group_data_for_json


def create_json_url_put_obj(group_data, group_object_data, override_object=False):
    group_data_for_json = {}
    if override_object:
        pass
    else:
        new_objects = list()
        for item in group_object_data:
            new_objects.append(item)
        for element in cfg.all_obj_domain[group_data['domain_name']]['urlgroups'][group_data['name']]['objects']:
            new_objects.append(element)
        group_data_for_json.update({"objects": new_objects})
        group_data_for_json.update({"type": "UrlGroup"})
        group_data_for_json.update({"name": group_data['name']})
        object_name, object_data = get_object_data(
            group_data['name'], group_data['domain_name'])
        group_data_for_json.update({"id": object_data['id']})
    return group_data_for_json

def create_json_group_obj(group_data, group_object_data, override_object=False):
    if all([group_data, group_object_data]):
        group_data_for_json = {}
        group_data_for_json.update({"objects": group_object_data})
        group_data_for_json.update({"type": "NetworkGroup"})
        group_data_for_json.update({"name": group_data['name']})
        return group_data_for_json


def create_json_urlgroup_obj(group_data, group_object_data, override_object=False):
    if all([group_data, group_object_data]):
        group_data_for_json = {}
        group_data_for_json.update({"objects": group_object_data})
        group_data_for_json.update({"type": "UrlGroup"})
        group_data_for_json.update({"name": group_data['name']})
        return group_data_for_json

def remove_space_from_name(group_name):
    if group_name:
        group_name = group_name.strip()
        if match_space(group_name):
            print(f'SPACE is present in group name {group_name}')
            # remove_space = do_ask_user_input('Remove SPACE from the group name [y/N] ') == 'y'
            # if remove_space:
            #     while (match_space(group_name)):
            #         group_name = remove_space_obj(group_name)
            while (match_space(group_name)):
                group_name = remove_space_obj(group_name)
    return group_name


def get_objects_data_for_group(object_data):
    tmp_dict = dict()
    if object_data:
        tmp_dict.update({'id': object_data.get('id'), 'type': object_data.get('type'), 'name': object_data.get('name') })
    return tmp_dict


def objects_non_override_to_group(domain_name, ws):
    """
    read xlsx, create groups of objects

    :param domain_name: domain name for each domain in FMC
    :type domain_name: str
    :param ws: Excel sheet name (should be linked with domain name)
    :type ws: openpyxl class object
    """
    # logging.info(f"\n{50*'#'}"
    #              f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to add Non-override groups\n'
    #              f"{50*'#'}")

    domains_add_groupobj = dict()
    domains_del_groupobj = dict()
    domains_add_override_group = dict()

    groupObject = dict()
    try:
        uuid = get_domain_uuid(domain_name)['uuid']
    except TypeError as error:
        logging.info(f'Domain {domain_name} do not exist either on FMC, either in Excel sheet')
        errors_filename = 'outputs/errors.txt'
        with open(errors_filename, "a") as f:
            f.write(f'Domain {domain_name} do not exist either on FMC, either in Excel sheet\n Error: {error}\n')

    ''' get size of xlsx sheet objects'''
    max_row = ws.max_row
    max_column = ws.max_column
    
    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1      
            

    add_group_payload = list()
    add_modify_group_payload = list()
    del_group_payload = list()

    # logging.info(f"\n{50*'#'}"
    #              f'\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to add NON override groups to {domain_name}\n'
    #              f"{50*'#'}")
    
    """
    Open excel sheet data as pandas DataFrame. Remove first row out of the DataFrame as data row. Use first row as columns name for the DataFrame.
    """    
    df = pd.DataFrame(ws.values)
    new_header = df.iloc[0]
    df = df[1:]
    df = df.rename(columns=new_header)
    
    """ temp list to include all data exported out of the DataFrame """
    list_sorted_groups_False_Add = list()
    
    df_groups_sort = df.groupby(['object_group_name','action'])
    
    for group, group_data1 in df_groups_sort:
        if (group[1] == 'add'):
            temp_group = df_groups_sort.get_group(group)
            temp = temp_group.to_dict('records')
            list_sorted_groups_False_Add.append(temp)
            
    for groups_desc in list_sorted_groups_False_Add:

        check_group_name = groups_desc[0].get('object_group_name')
        check_object_type = groups_desc[0].get('type')
        group_data = dict()
        group_object_data = list()
        
        if not check_if_object_already_exist(check_group_name, domain_name):

            for obj in groups_desc:   
                group_name = obj.get('object_group_name')
                group_name = remove_space_from_name(group_name)
                obj_name = obj.get('object')
                obj_name = remove_space_from_name(obj_name)

                if not check_if_object_already_exist(obj_name, domain_name):
                    logging.info(
                        f'Error: object {obj_name} for group {check_group_name} do NOT exist!')
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(
                            f'Object {obj_name} do not exist\n')
                    continue
                
                '''So, new line has new group. Thus we need:
                1) Record objects and data for the previous group
                2) Don't drop new group data and new object and record them to group_pre_data'''
                
                try:
                    object_name, object_data = get_object_data(obj_name, domain_name)
                    object_data_for_group = get_objects_data_for_group(object_data)
                    group_object_data.append(object_data_for_group)
                except KeyError as error:
                    logging.info(f'Object {obj_name} do not exist')
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(f'Object {obj_name} do not exist\n Error: {error}\n')
                
                group_data.update({'name': group_name})
                
                group_data.update({'domain_name': domain_name})
                group_data.update({'objects': group_object_data})

            if group_object_data:
                '''Create json-like structure (using dict) for API request'''
                group_objects_json = create_json_group_obj(
                    group_data,
                    group_object_data)
                    
                
                '''add to common list which would be run by API request later'''
                add_group_payload.append(
                    {group_name: group_objects_json})


        elif check_if_object_already_exist(check_group_name, domain_name):

            for obj in groups_desc:
                group_name = obj.get('object_group_name')
                group_name = remove_space_from_name(group_name)
                obj_name = obj.get('object')
                obj_name = remove_space_from_name(obj_name)
            
                if not check_if_object_already_exist(obj_name, domain_name):
                    f'Error: object {obj_name} for group {check_group_name} do NOT exist!'
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(
                            f'Object {obj_name} DO not exist\n')
                    continue
            
                '''So, new line has new group. Thus we need:
                1) Record objects and data for the previous group
                2) Don't drop new group data and new object and record them to group_pre_data'''

                try:
                    object_name, object_data = get_object_data(obj_name, domain_name)
                    object_data_for_group = get_objects_data_for_group(object_data)
                    group_object_data.append(object_data_for_group)
                except KeyError as error:
                    logging.info(f'Object {obj_name} do not exist')
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(f'Object {obj_name} do not exist\n Error: {error}\n')
                
                group_data.update({'name': group_name})
               
                group_data.update({'domain_name': domain_name})
                group_data.update({'objects': group_object_data})

            if group_object_data:
                '''Create json-like structure (using dict) for API request'''
                group_objects_json = create_json_put_obj(
                    group_data,
                    group_object_data)
                
                '''add to common list which would be run by API request later'''
                add_modify_group_payload.append(
                    {group_name: group_objects_json})                

    if add_group_payload:
        '''Merge same group data from multiple dictionaries from add_group_payload list into one dictionary with groups values list '''
        all_groups_dict = dict()
        for group_data in add_group_payload:
            lst1 = list()
            for group_name, group_value in group_data.items():
                if group_name in all_groups_dict:
                    for element in all_groups_dict[group_name]:
                        lst1.append(element)
                    lst1.append(group_data[group_name])
                    all_groups_dict.update({group_name: lst1})
                else:
                    lst1.append(group_data[group_name])
                    all_groups_dict.update({group_name: lst1})
        ''' Merge same group data into one dictionary for API request '''
        all_in_one_dict = dict()
        for group, group_lists in all_groups_dict.items():
            group_objects = list()
            group_data_dict = dict()
            for item in group_lists:
                group_objects += item.get('objects')
                for key, value in item.items():
                    group_data_dict.update({key: value})
            group_data_dict.update({'objects': group_objects})
            all_in_one_dict.update({group: group_data_dict})

        payload_list = []
        for group, group_data in all_in_one_dict.items():
            payload_list.append(group_data)
        
        domains_add_groupobj.update(
            {domain_name: {'groupObject': payload_list}})
    if del_group_payload:
        domains_del_groupobj.update(
            {domain_name: {'groupObject': del_group_payload}})

    if add_modify_group_payload:
        ''' now to PUT our list of group objects '''
        for item in add_modify_group_payload:
            for group, group_data in item.items():
                try:
                    if group_data:
                        put_networkgroups(group_data, domain_name)
                except KeyError:
                    logging.info(
                        f'domain = {domain_name} do not has Group objects to add')
    if domains_add_groupobj:
        ''' now to POST our list of group objects '''
        for domain, domain_data in domains_add_groupobj.items():
            try:
                if domain_data:
                    post_groups_objects(
                        domain_data['groupObject'],
                        get_domain_uuid(domain)['uuid'])
            except KeyError:
                logging.info(
                    f'domain = {domain} do not has Group objects to add')
        # logging.info(f"\n{50*'#'}"
        #              f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished adding Non-override groups\n'
        #              f"{50*'#'}")
    if any([domains_add_groupobj, domains_add_override_group]):
        update_all_objects(domain_name)
        update_all_networkgroups(domain_name)


def url_objects_groups(domain_name, ws):
    """
    read xlsx, create groups of objects

    :param domain_name: domain name for each domain in FMC
    :type domain_name: str
    :param ws: Excel sheet name (should be linked with domain name)
    :type ws: openpyxl class object
    """
    domains_add_groupobj = dict()
    domains_del_groupobj = dict()

    try:
        uuid = get_domain_uuid(domain_name)['uuid']
    except TypeError as error:
        logging.info(
            f'Domain {domain_name} do not exist either on FMC, either in Excel sheet')
        errors_filename = 'outputs/errors.txt'
        with open(errors_filename, "a") as f:
            f.write(
                f'Domain {domain_name} do not exist either on FMC, either in Excel sheet\n Error: {error}\n')

    ''' get size of xlsx sheet objects'''
    max_row = ws.max_row
    max_column = ws.max_column

    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1

    add_group_payload = list()
    add_modify_group_payload = list()
    del_group_payload = list()

    # logging.info(f"\n{50*'#'}"
    #              f'\n{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to add NON override groups to {domain_name}\n'
    #              f"{50*'#'}")

    """
    Open excel sheet data as pandas DataFrame. Remove first row out of the DataFrame as data row. Use first row as columns name for the DataFrame.
    """
    df = pd.DataFrame(ws.values)
    new_header = df.iloc[0]
    df = df[1:]
    df = df.rename(columns=new_header)

    """ temp list to include all data exported out of the DataFrame """
    list_sorted_groups_False_Add = list()

    df_groups_sort = df.groupby(['url_group_name', 'action'])

    for group, group_data1 in df_groups_sort:
        if (group[1] == 'add'):
            temp_group = df_groups_sort.get_group(group)
            temp = temp_group.to_dict('records')
            list_sorted_groups_False_Add.append(temp)

    for groups_desc in list_sorted_groups_False_Add:

        check_group_name = groups_desc[0].get('url_group_name')
        group_data = dict()
        group_object_data = list()

        if not check_if_object_already_exist(check_group_name, domain_name):

            for obj in groups_desc:
                group_name = obj.get('url_group_name')
                group_name = remove_space_from_name(group_name)
                obj_name = obj.get('url')
                obj_name = remove_space_from_name(obj_name)

                if not check_if_object_already_exist(obj_name, domain_name):
                    logging.info(f'Error: url {obj_name} do NOT exist!')
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(
                            f'Object {obj_name} do not exist\n')
                    continue

                '''So, new line has new group. Thus we need:
                1) Record objects and data for the previous group
                2) Don't drop new group data and new object and record them to group_pre_data'''

                try:
                    object_name, object_data = get_object_data(
                        obj_name, domain_name)
                    object_data_for_group = get_objects_data_for_group(
                        object_data)
                    group_object_data.append(object_data_for_group)
                except KeyError as error:
                    logging.info(f'url {obj_name} do not exist')
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(
                            f'Object {obj_name} do not exist\n Error: {error}\n')

                group_data.update({'name': group_name})

                group_data.update({'domain_name': domain_name})
                group_data.update({'objects': group_object_data})

            if group_object_data:
                '''Create json-like structure (using dict) for API request'''
                group_objects_json = create_json_urlgroup_obj(
                    group_data,
                    group_object_data)

                '''add to common list which would be run by API request later'''
                add_group_payload.append(
                    {group_name: group_objects_json})

        elif check_if_object_already_exist(check_group_name, domain_name):

            for obj in groups_desc:
                group_name = obj.get('url_group_name')
                group_name = remove_space_from_name(group_name)
                obj_name = obj.get('url')
                obj_name = remove_space_from_name(obj_name)

                if not check_if_object_already_exist(obj_name, domain_name):
                    f'Error: object {obj_name} for group {check_group_name} do NOT exist!'
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(
                            f'Object {obj_name} DO not exist\n')
                    continue

                '''So, new line has new group. Thus we need:
                1) Record objects and data for the previous group
                2) Don't drop new group data and new object and record them to group_pre_data'''

                try:
                    object_name, object_data = get_object_data(
                        obj_name, domain_name)
                    object_data_for_group = get_objects_data_for_group(
                        object_data)
                    group_object_data.append(object_data_for_group)
                except KeyError as error:
                    logging.info(f'Object {obj_name} do not exist')
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(
                            f'Object {obj_name} do not exist\n Error: {error}\n')

                group_data.update({'name': group_name})

                group_data.update({'domain_name': domain_name})
                group_data.update({'objects': group_object_data})

            if group_object_data:
                '''Create json-like structure (using dict) for API request'''
                group_objects_json = create_json_url_put_obj(
                    group_data,
                    group_object_data)

                '''add to common list which would be run by API request later'''
                add_modify_group_payload.append(
                    {group_name: group_objects_json})

    if add_group_payload:
        '''Merge same group data from multiple dictionaries from add_group_payload list into one dictionary with groups values list '''
        all_groups_dict = dict()
        for group_data in add_group_payload:
            lst1 = list()
            for group_name, group_value in group_data.items():
                if group_name in all_groups_dict:
                    for element in all_groups_dict[group_name]:
                        lst1.append(element)
                    lst1.append(group_data[group_name])
                    all_groups_dict.update({group_name: lst1})
                else:
                    lst1.append(group_data[group_name])
                    all_groups_dict.update({group_name: lst1})
        ''' Merge same group data into one dictionary for API request '''
        all_in_one_dict = dict()
        for group, group_lists in all_groups_dict.items():
            group_objects = list()
            group_data_dict = dict()
            for item in group_lists:
                group_objects += item.get('objects')
                for key, value in item.items():
                    group_data_dict.update({key: value})
            group_data_dict.update({'objects': group_objects})
            all_in_one_dict.update({group: group_data_dict})

        payload_list = []
        for group, group_data in all_in_one_dict.items():
            payload_list.append(group_data)

        domains_add_groupobj.update(
            {domain_name: {'groupObject': payload_list}})
    if del_group_payload:
        domains_del_groupobj.update(
            {domain_name: {'groupObject': del_group_payload}})

    if add_modify_group_payload:
        ''' now to PUT our list of group objects '''
        for item in add_modify_group_payload:
            for group, group_data in item.items():
                try:
                    if group_data:
                        put_urlgroups(group_data, domain_name)
                except KeyError:
                    logging.info(
                        f'domain = {domain_name} do not has UrlGroup objects to add')
    if domains_add_groupobj:
        ''' now to POST our list of group objects '''
        for domain, domain_data in domains_add_groupobj.items():
            try:
                if domain_data:
                    post_urlgroups_objects(
                        domain_data['groupObject'],
                        get_domain_uuid(domain)['uuid'])
            except KeyError:
                logging.info(
                    f'domain = {domain} do not has UrlGroup objects to add')
        # logging.info(f"\n{50*'#'}"
        #              f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished adding Non-override groups\n'
        #              f"{50*'#'}")
    if any([domains_add_groupobj, add_modify_group_payload, add_group_payload]):
        update_all_objects(domain_name)


def check_xlsx_sheet_empty_lines(ws, max_row):
    ''' check for empty lines in Excel '''
    max_value = copy.deepcopy(max_row)
    for i in reversed(range(1, max_value+1)):
        some_data = ws.cell(row=i, column=1).value
        if some_data:
            max_row = i
            break
        elif not some_data:
            max_row -= 1
    return max_row


def update_all_objects(domain_name):
    """
    update_all_objects update all_jbjects with uuid, name, description, links to self

    :param domain_name: domain_name (function shoudl be run for each domain)
    :type domain_name: str
    :return: return all updated objects, return data shoudl be updated to all_objects global dictionary
    :rtype: dict()
    """       
    # domain_obj_domain, domain_ids = get_all_objects_for_domain(domain_name)
    domain_obj_domain, domain_ids = get_all_objects_for_domain_no_check_ids(domain_name)
    merged_all_obj_domain = dict()
    for domain_name, objects_data in cfg.all_obj_domain.items():
        tmp_hosts = dict()
        tmp_ranges = dict()
        tmp_networks = dict()
        tmp_urls = dict()
        tmp_networkgroups = dict()
        tmp_urlgroups = dict()
        
        if domain_obj_domain.get(domain_name):
            if domain_obj_domain.get(domain_name).get('hosts'):
                tmp_hosts = {**objects_data.get('hosts'), **domain_obj_domain.get(domain_name).get('hosts')}
            else:
                tmp_hosts = objects_data.get('hosts')
        else:
            tmp_hosts = objects_data.get('hosts')
        if domain_obj_domain.get(domain_name):
            if domain_obj_domain.get(domain_name).get('ranges'):
                tmp_ranges = {**objects_data.get('ranges'), **domain_obj_domain.get(domain_name).get('ranges')}
            else:
                tmp_ranges = objects_data.get('ranges')
        else:
            tmp_ranges = objects_data.get('ranges')
        if domain_obj_domain.get(domain_name):
            if domain_obj_domain.get(domain_name).get('networks'):
                tmp_networks = {**objects_data.get('networks'), **domain_obj_domain.get(domain_name).get('networks')}
            else:
                tmp_networks = objects_data.get('networks')
        else:
            tmp_networks = objects_data.get('networks')
        if domain_obj_domain.get(domain_name):
            if domain_obj_domain.get(domain_name).get('urls'):
                tmp_urls = {**objects_data.get('urls'), **domain_obj_domain.get(domain_name).get('urls')}
            else:
                tmp_urls = objects_data.get('urls')
        else:
            tmp_urls = objects_data.get('urls')
        if domain_obj_domain.get(domain_name):
            if domain_obj_domain.get(domain_name).get('networkgroups'):
                tmp_networkgroups = {**objects_data.get('networkgroups'), **domain_obj_domain.get(domain_name).get('networkgroups')}
            else:
                tmp_networkgroups = objects_data.get('networkgroups')
        else:
            tmp_networkgroups = objects_data.get('networkgroups')
        if domain_obj_domain.get(domain_name):
            if domain_obj_domain.get(domain_name).get('urlgroups'):
                tmp_urlgroups = {**objects_data.get('urlgroups'), **domain_obj_domain.get(domain_name).get('urlgroups')}
            else:
                tmp_urlgroups = objects_data.get('urlgroups')
        else:
            tmp_urlgroups = objects_data.get('urlgroups')
        merged_all_obj_domain.update({domain_name: {'hosts': tmp_hosts, 'ranges': tmp_ranges, 'networks': tmp_networks, 'urls': tmp_urls, 'networkgroups': tmp_networkgroups, 'urlgroups': tmp_urlgroups}})

    merged_all_ids_domain = dict()
    if domain_ids:
        for domain_name, object_id_dicts in cfg.all_ids_domain.items():
            tmp_dict = dict()
            if all([object_id_dicts, domain_ids.get(domain_name)]):
                tmp_dict = { **object_id_dicts, **domain_ids.get(domain_name)}
            elif all([object_id_dicts, not domain_ids.get(domain_name)]):
                tmp_dict = object_id_dicts
            elif all([not object_id_dicts, domain_ids.get(domain_name)]):
                tmp_dict = domain_ids.get(domain_name)
            merged_all_ids_domain.update({domain_name: tmp_dict})
    cfg.all_obj_domain = merged_all_obj_domain
    cfg.all_ids_domain = merged_all_ids_domain


def update_all_networkgroups(domain_name):
    """
    update_all_networkgroups update all_networkgroups

    :param domain_name: domain_name (function shoudl be run for each domain)
    :type domain_name: str
    :return: return all updated objects, return data shoudl be updated to all_objects global dictionary
    :rtype: dict()
    """
    domain_detailed_networkgroups = get_all_detailed_groups_for_domain(domain_name)

    merged_detailed_networkgroups_domain = dict()
    for domain_name, objects_data in cfg.all_detailed_networkgroups.items():
        tmp_networkgroups = dict()
        if domain_detailed_networkgroups.get(domain_name):
            tmp_networkgroups = {**domain_detailed_networkgroups.get(domain_name), **objects_data}
        else:
            tmp_networkgroups = cfg.all_detailed_networkgroups.get(domain_name)
        merged_detailed_networkgroups_domain.update({domain_name: tmp_networkgroups})

    cfg.all_detailed_networkgroups = merged_detailed_networkgroups_domain
 

def sort_sheets_by_creation(all_sheets):
    """
    Create 4 lists to sort xlsx data by order:
    first: create objects
    second: create groups
    third: assing objects into groups
    """
    sheet_process_order = []
    group_list = []
    objects_list = []
    objects_into_groups_list = []
    for sheet in all_sheets:
        names = sheet.strip().split('.')       
        if 'groups' in names:
            objects_into_groups_list.append(sheet)
        else:
            objects_list.append(sheet)
    sheet_process_order = objects_list + group_list + objects_into_groups_list
    return sheet_process_order


if __name__ == "__main__":
    cfg.init()
    output_dir = 'outputs'
    try:
        os.makedirs(output_dir)
    except OSError as e:
        pass 
    errors_filename = 'outputs/errors.txt'
    with open(errors_filename, "w") as f:
        f.write('')
    
    cfg.all_domains_json = get_all_domains_data()
    cfg.all_obj_domain, cfg.all_ids_domain = get_all_objects_with_domains()
    cfg.all_devices = get_all_devices()
    cfg.all_detailed_networkgroups = get_all_groups_info()
    
    if cfg.check_if_file_exist(cfg.input_xlsx):  
        ''' read and parse data out of the XLSX Commutation map '''
        wb = Workbook()
        wb = load_workbook(cfg.input_xlsx, read_only=False)
        all_sheets = wb.sheetnames

        sorted_sheets = sort_sheets_by_creation(all_sheets)
        cfg.sorted_sheets = sorted_sheets
        
        # diff_before_filename = f'{cfg.diff_before_filename}-{datetime.now().strftime("%Y-%m-%d-%H-%M-%S")}.xlsx'
        # create_diff_excel_file(diff_before_filename)
        
        for sheet in sorted_sheets:
            ws = wb[sheet]
            names = sheet.strip().split('.')
            if not 'groups' in names and not 'urlgrps' in names:
                domain_name = sheet.replace('.','/')
                object_change_function(domain_name, ws)
            elif 'groups' in names:
                if names[-1] == 'groups':
                    names.remove(names[-1])
                domain_name = '/'.join(names)
                del_group_from_parent_group(domain_name, ws)
                change_hosts_nets_in_parent_group(domain_name, ws)
                objects_non_override_to_group(domain_name, ws)
            elif 'urlgrps' in names:
                if names[-1] == 'urlgrps':
                    names.remove(names[-1])
                domain_name = '/'.join(names)
                change_urls_in_group(domain_name, ws)
                url_objects_groups(domain_name, ws)
        for sheet in sorted_sheets:
            ws = wb[sheet]
            names = sheet.strip().split('.')
            if not 'groups' in names and not 'urlgrps' in names:
                domain_name = sheet.replace('.', '/')
                del_hosts_nets_for_obj_sheet(domain_name, ws)
        wb.close()
            
        create_xlsx_and_sheets(cfg.output_xlsx)
        write_hosts_network_objects_to_xlsx(cfg.output_xlsx)
        write_group_objects_to_xlsx(cfg.output_xlsx)
        write_urlgrps_to_xlsx(cfg.output_xlsx)
        # diff_after_filename = f'{cfg.diff_after_filename}-{datetime.now().strftime("%Y-%m-%d-%H-%M-%S")}.xlsx'
        # create_diff_excel_file(diff_after_filename)
        logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Execution has been completed with no major exceptions. Done.\n')
    else:
        create_xlsx_and_sheets(cfg.input_xlsx)
        write_hosts_network_objects_to_xlsx(cfg.input_xlsx)
        write_group_objects_to_xlsx(cfg.input_xlsx)
        write_urlgrps_to_xlsx(cfg.input_xlsx)