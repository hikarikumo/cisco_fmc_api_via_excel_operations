#!/usr/bin/python3
'''
Create Excel xlsx file with objects and groups information in different domains
'''
from get_func import get_all_domains_data, get_all_objects_with_domains, get_all_groups_info, get_all_devices
import cfg
from datetime import datetime
from string import ascii_uppercase
from openpyxl.styles import Font # Connect styles for text
from openpyxl import Workbook
from openpyxl import load_workbook
import logging
import os


logging.basicConfig(format='%(threadName)s %(name)s %(levelname)s: %(message)s',level=logging.INFO)


def create_xlsx_and_sheets(output_xlsx):
    domains_list = list()
    for domain_data in cfg.all_domains_json:
        domain_name = domain_data['name']
        domain_name = domain_name.replace('/', '.')
        domains_list.append(domain_name)
    
    sorted_domains_list = sorted(domains_list, key=len)
    
    wb = Workbook()
    wb.active
    
    for domain in sorted_domains_list:
        wb.create_sheet(f'{domain}')
        wb.create_sheet(f'{domain}.groups')
        wb.create_sheet(f'{domain}.urlgrps')
    wb.remove(wb.active)
    wb.save(output_xlsx)
    wb.close()
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Domain tabs were created and were written to {output_xlsx} \n')


def get_sheet_name_for_domain(domain_name, all_sheets):
    domain_for_sheet = domain_name.replace('/', '.')
    for sheet in all_sheets:
        if sheet == domain_for_sheet:
            return sheet


def write_hosts_network_objects_to_xlsx(output_xlsx):
    system_hosts = cfg.system_hosts
    system_networks = cfg.system_networks
    
    wb = Workbook()
    wb = load_workbook(output_xlsx, read_only=False)
    all_sheets = wb.sheetnames
    
    for domain, objects in cfg.all_obj_domain.items():
        ws=wb[get_sheet_name_for_domain(domain, all_sheets)]
        index = 1
        ws.cell(row=1, column=1, value='object_name').font = Font(b=True)
        ws.cell(row=1, column=2, value='object').font = Font(b=True)
        ws.cell(row=1, column=3, value='action').font = Font(b=True)
        ws.cell(row=1, column=4, value='type').font = Font(b=True)
        for column in ascii_uppercase:
            if column == 'A': 
                ws.column_dimensions[column].width = 50
            elif column == 'B':
                ws.column_dimensions[column].width = 50
            elif column == 'C':
                ws.column_dimensions[column].width = 10
            elif column == 'D':
                ws.column_dimensions[column].width = 15
        if objects.get('hosts'):
            for obj_key, obj_value in objects.get('hosts').items():
                if obj_key in system_hosts:
                    continue
                ws.cell(row=index + 1, column=1, value=obj_key)
                ws.cell(row=index + 1, column=2, value=obj_value.get('value'))
                ws.cell(row=index + 1, column=4, value=obj_value.get('type'))
                for j in range(1, 4):
                    ws.cell(row=index + 1, column=j).font = Font(color='008080')
                index += 1
        if objects.get('ranges'):
            for obj_key, obj_value in objects.get('ranges').items():
                ws.cell(row=index + 1, column=1,value=obj_key)
                ws.cell(row=index + 1, column=2,value=obj_value.get('value'))
                ws.cell(row=index + 1, column=4, value=obj_value.get('type'))
                for j in range(1, 4):
                    ws.cell(row=index + 1,
                            column=j).font = Font(color='008080')
                index += 1
        if objects.get('networks'):
            for obj_key, obj_value in objects.get('networks').items():
                if obj_key in system_networks:
                    continue
                ws.cell(row=index + 1, column=1, value=obj_key)
                ws.cell(row=index + 1, column=2, value=obj_value.get('value'))
                ws.cell(row=index + 1, column=4, value=obj_value.get('type'))
                for j in range(1, 4):
                    ws.cell(row=index + 1, column=j).font = Font(color='008080')
                index += 1
        if objects.get('urls'):
            for obj_key, obj_value in objects.get('urls').items():
                if obj_key in system_networks:
                    continue
                ws.cell(row=index + 1, column=1, value=obj_key)
                ws.cell(row=index + 1, column=2, value=obj_value.get('url'))
                ws.cell(row=index + 1, column=4, value=obj_value.get('type'))
                for j in range(1, 4):
                    ws.cell(row=index + 1, column=j).font = Font(color='008080')
                index += 1
    wb.save(output_xlsx)
    wb.close()
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Hosts, Ranges, Networks were written to {output_xlsx} \n')


def write_group_objects_to_xlsx(output_xlsx):
    wb = Workbook()
    wb = load_workbook(output_xlsx, read_only=False)
    all_sheets = wb.sheetnames

    for domain, networkgroups in cfg.all_detailed_networkgroups.items():
        ws = wb[f'{get_sheet_name_for_domain(domain, all_sheets)}.groups']
        index = 1
        ws.cell(row=1, column=1, value='object_group_name').font = Font(b=True)
        ws.cell(row=1, column=2, value='object').font = Font(b=True)
        ws.cell(row=1, column=3, value='action').font = Font(b=True)
        ws.cell(row=1, column=4, value='type').font = Font(b=True)
        for column in ascii_uppercase:
            if column == 'A':
                ws.column_dimensions[column].width = 50
            elif column == 'B':
                ws.column_dimensions[column].width = 50
            elif column == 'C':
                ws.column_dimensions[column].width = 10
            elif column == 'D':
                ws.column_dimensions[column].width = 15
        for group_name, group_value in networkgroups.items():
            if group_name == 'IPv4-Private-All-RFC1918' or group_name == 'any':
                continue
            if group_value.get('objects'):
                for item in group_value['objects']:
                    ws.cell(row=index + 1, column=1,
                            value=group_value.get('name'))
                    ws.cell(row=index + 1, column=2,
                            value=item.get('name'))
                    ws.cell(row=index + 1, column=4,
                            value=group_value.get('type'))
                    for j in range(1, 4):
                        ws.cell(row=index + 1,
                                column=j).font = Font(color='008080')
                    index += 1
    wb.save(output_xlsx)
    wb.close()
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Groups were written to {output_xlsx} \n')


def write_urlgrps_to_xlsx(output_xlsx):
    wb = Workbook()
    wb = load_workbook(output_xlsx, read_only=False)
    all_sheets = wb.sheetnames

    for domain, object_data in cfg.all_obj_domain.items():
        if object_data.get('urlgroups'):
            ws = wb[f'{get_sheet_name_for_domain(domain, all_sheets)}.urlgrps']
            index = 1
            ws.cell(row=1, column=1, value='url_group_name').font = Font(b=True)
            ws.cell(row=1, column=2, value='url').font = Font(b=True)
            ws.cell(row=1, column=3, value='action').font = Font(b=True)
            ws.cell(row=1, column=4, value='type').font = Font(b=True)
            for column in ascii_uppercase:
                if column == 'A':
                    ws.column_dimensions[column].width = 50
                elif column == 'B':
                    ws.column_dimensions[column].width = 50
                elif column == 'C':
                    ws.column_dimensions[column].width = 10
                elif column == 'D':
                    ws.column_dimensions[column].width = 15
            for urlgrp_name, urlgrp_value in object_data.get('urlgroups').items():
                if urlgrp_value.get('objects'):
                    for item in urlgrp_value['objects']:
                        ws.cell(row=index + 1, column=1,
                                value=urlgrp_name)
                        ws.cell(row=index + 1, column=2,
                                value=item.get('name'))
                        ws.cell(row=index + 1, column=4,
                                value=urlgrp_value.get('type'))
                        for j in range(1, 4):
                            ws.cell(row=index + 1,
                                    column=j).font = Font(color='008080')
                        index += 1
        else:
            sheet_for_deletion = wb[f'{get_sheet_name_for_domain(domain, all_sheets)}.urlgrps']
            wb.remove(sheet_for_deletion)
            
    wb.save(output_xlsx)
    wb.close()
    logging.info(
        f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} UrlGroups were written to {output_xlsx} \n')


if __name__ == "__main__":
    cfg.init()

    output_dir = 'outputs'
    try:
        os.makedirs(output_dir)
    except OSError as e:
        pass

    errors_filename = 'outputs/errors.txt'
    with open(errors_filename, "w") as f:
        f.write('')

    # input_xlsx = 'FMC_VFHU_objects.xlsx'
    # output_xlsx = 'FMC_VFHU_downloaded_objects.xlsx'

    cfg.all_domains_json = get_all_domains_data()
    cfg.all_obj_domain, cfg.all_ids_domain = get_all_objects_with_domains()
    cfg.all_devices = get_all_devices()
    cfg.all_detailed_networkgroups = get_all_groups_info()
    
    create_xlsx_and_sheets(cfg.output_xlsx)
    write_hosts_network_objects_to_xlsx(cfg.output_xlsx)
    write_group_objects_to_xlsx(cfg.output_xlsx)
    write_urlgrps_to_xlsx(cfg.output_xlsx)
    logging.info(f"\n{50*'#'}"
                 f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Execution has been completed with no major exceptions. Done.\n'
                 f"{50*'#'}")
