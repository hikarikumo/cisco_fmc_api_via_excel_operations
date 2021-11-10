#!/usr/bin/python3
'''
GET information about Security Zones (SZ), portobjects, TCP and UDP objects on FMC
GET ACP ACL rules
Download new ACP and ACL from the FMC
'''
from datetime import datetime
from get_func import get_all_domains_data, get_all_objects_with_domains, get_all_objects_for_domain, get_all_groups_info, get_all_devices, get_domain_uuid, get_object_data, check_if_object_already_exist, check_object_type, get_all_detailed_groups_for_domain, check_object_type, get_domain_uuid, get_domain_name_by_uuid, api_call_counter, check_token_status
from string import ascii_uppercase
from pathlib import Path
from openpyxl.styles import Font # Connect styles for text
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import json, yaml
import requests
import logging
import re
import os
import cfg


logging.basicConfig(
    format='%(threadName)s %(name)s %(levelname)s: %(message)s',
    level=logging.INFO)


def get_acp_sheet(all_sheets):
    """
    get_acp_sheet get sheet name which has data for ACP, ACL creation

    :param all_sheets: all_xslsx sheet names. Multiple sheets contain objects, groups, overrides, ACP lists data
    :type all_sheets: string
    :return: sheet name which has data for ACP lists
    :rtype: string
    """
    acp_sheets = []
    for sheet in all_sheets:
        names = sheet.strip().split('.')
        if 'ACP' in names:
            acp_sheets.append(sheet)
    return acp_sheets


def get_all_sz():
    """
    get_all_sz create dictionary with all security zones to be used as the reference

    :return: dictionary with all sz names as keys
    :rtype: dict()
    """
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to collect all Security Zones data\n')
    objects_dict = {}
    all_sz = {}
    all_sz_domain = {}
    
    offset_value = 0
    limit_value = 1000
    obj_types = ['securityzones']
    for domain in cfg.all_domains_json:
        domain_all_sz = {}

        for obj_type in obj_types:
            objects_data = []
            check_token_status()
            api_call_counter()
            resp = requests.get(
                f"https://{cfg.fmc_ip}/api/fmc_config/v1/domain/{get_domain_uuid(domain['name'])['uuid']}/object/{obj_type}?offset={offset_value}&limit={limit_value}",
                headers=cfg.headers_json,
                verify=False)
            logging.info(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} get {obj_type} for {domain["name"]} status code: {str(resp.status_code)}\n')
            if str(resp.status_code).startswith('2'):
                try:
                    objects_data = json.loads(resp.text)['items']
                except KeyError:
                    logging.info(
                        f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} No security zones received.\n Maybe none SZ configured?')
                try:
                    next_url = resp.json()['paging']['next'][0]
                except KeyError as error:
                    logging.info(
                        f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Reached end of pages for {obj_type} in {domain["name"]}')
                    next_url = ''
                while next_url:
                    check_token_status()
                    api_call_counter()
                    resp = requests.get(next_url,
                                        headers=cfg.headers_json,
                                        verify=False)
                    if str(resp.status_code).startswith('2'):
                        try:
                            objects_data += json.loads(resp.text)['items']
                        except KeyError:
                            logging.info(
                                f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} No security zones received.\n Maybe none SZ configured?')
                    try:
                        next_url = resp.json()['paging']['next'][0]
                    except KeyError:
                        logging.info(
                            f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Reached end of pages for {obj_type} in {domain["name"]}')
                        next_url = ''

                for obj in objects_data:
                    if obj['name'] in list(all_sz):
                        continue
                    else:
                        domain_all_sz.update({obj['name']: obj})
                        all_sz.update({obj['name']: obj})

            else:
                logging.info(
                    f' {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Object request has returned incorrect status code {resp.status_code}')
                continue
        all_sz_domain.update({domain['name']: domain_all_sz})
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished collection of Security Zones\n')
    return all_sz, all_sz_domain


def get_all_portobjects():
    """
    get_all_portobjects create dictionary with all security zones to be used as the reference

    :return: dictionary with all portobject names as keys
    :rtype: dict()
    """
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to collect all PortObjects data\n')
    objects_dict = {}
    all_portobjects = {}
    all_portobjects_domain = {}
    
    offset_value = 0
    limit_value = 1000
    obj_types = ['protocolportobjects']
    for domain in cfg.all_domains_json:
        domain_all_portobjects = {}

        for obj_type in obj_types:
            objects_data = []
            check_token_status()
            api_call_counter()
            resp = requests.get(
                f"https://{cfg.fmc_ip}/api/fmc_config/v1/domain/{get_domain_uuid(domain['name'])['uuid']}/object/{obj_type}?offset={offset_value}&limit={limit_value}",
                headers=cfg.headers_json,
                verify=False)
            logging.info(
                f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} get {obj_type} for {domain["name"]} status code: {str(resp.status_code)}\n')
            if str(resp.status_code).startswith('2'):
                try:
                    objects_data = json.loads(resp.text)['items']
                except KeyError:
                    logging.info(
                        f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} No PortObjects received.\n Maybe none PortObjects configured?')
                try:
                    next_url = resp.json()['paging']['next'][0]
                except KeyError as error:
                    logging.info(
                        f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Reached end of pages for {obj_type} in {domain["name"]}')
                    next_url = ''
                while next_url:
                    check_token_status()
                    api_call_counter()
                    resp = requests.get(next_url,
                                        headers=cfg.headers_json,
                                        verify=False)
                    # time.sleep(0.7)
                    if str(resp.status_code).startswith('2'):
                        try:
                            objects_data += json.loads(resp.text)['items']
                        except KeyError:
                            logging.info(
                                f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} No PortObjects received.\n Maybe none PortObjects configured?')

                    try:
                        next_url = resp.json()['paging']['next'][0]
                    except KeyError as error:
                        logging.info(
                            f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Reached end of pages for {obj_type} in {domain["name"]}')
                        next_url = ''

                for obj in objects_data:
                    if obj['name'] in list(all_portobjects):
                        continue
                    else:
                        domain_all_portobjects.update({obj['name']: obj})
                        all_portobjects.update({obj['name']: obj})

            else:
                logging.info(
                    f' {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Object request has returned incorrect status code {resp.status_code}')
                continue
        all_portobjects_domain.update({domain['name']: domain_all_portobjects})

    # with open('outputs/all_portobjects.yaml', 'w') as dest, open('outputs/all_portobjects_domain.yaml', 'w') as dest2:
    #     yaml.SafeDumper.ignore_aliases = lambda *args: True
    #     yaml.safe_dump(all_portobjects, dest, default_flow_style=False)
    #     yaml.safe_dump(all_portobjects_domain, dest2, default_flow_style=False)
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished collection of Security Zones\n')
    return all_portobjects, all_portobjects_domain


def get_domain_from_sz(zone_name):
    domains_list = list()
    for domain in cfg.all_domains_json:
        domains_list.append(domain['name'])
    domains_list.sort(key=len)
    for domain in domains_list:
        if zone_name in domain:
            return domain


def get_domain_for_portobject(domain_name):
    domains_list = list()
    for domain in cfg.all_domains_json:
        domains_list.append(domain['name'])
    domains_list.sort(key=len)
    for domain in domains_list:
        if domain_name in domain:
            return domain


def sort_zones_by_domain(list_of_tuples):
    domain_zones = dict()
    for domain, zone in list_of_tuples:
        domain_zones.setdefault(domain, []).append(zone)
    return domain_zones


def sort_portobjects_by_domain(list_of_tuples):
    domain_portobjects = dict()
    for domain, portobject in list_of_tuples:
        domain_portobjects.setdefault(domain, []).append(portobject)
    return domain_portobjects


# def create_domains_list():
#     domains_list = list()
#     for domain in cfg.all_domains_json:
#         domains_list.append(domain['name'])
#     domains_list.sort(key=len)
    
    
def create_json_add_sz(domain_zones_dict):
    domain_sz_json = {}
    for domain, zones in domain_zones_dict.items():
        for zone in zones:
            sz_json_payload = dict()
            sz_json_payload.update({'name': zone})
            sz_json_payload.update({'type': 'SecurityZone'})
            sz_json_payload.update({'interfaceMode': 'ROUTED'})
            domain_sz_json.setdefault(domain, []).append(sz_json_payload)
    return domain_sz_json


def create_json_add_portobjects(domain_portobjects_dict):
    domain_portobjects_json = {}
    for domain, portobjects in domain_portobjects_dict.items():
        for portobject in portobjects:
            protocol = str()
            ports = str()
            match = re.search(
                r'(?P<protocol>.+?(?=\.))'
                r'(?:\.)'
                r'(?P<ports>\S+)', portobject)
            if portobject.lower() == 'icmp':
                continue
            if match:
                protocol = match.group('protocol').upper()
                ports = match.group('ports')
            portobject_json_payload = dict()
            portobject_json_payload.update({'name': portobject})
            portobject_json_payload.update({'type': 'ProtocolPortObject'})
            portobject_json_payload.update({'overridable': 'false'})
            portobject_json_payload.update({'protocol': protocol})
            portobject_json_payload.update({'port': ports})
            domain_portobjects_json.setdefault(
                domain, []).append(portobject_json_payload)
    return domain_portobjects_json


def post_sz_objects(sz_payload, domain_uuid):
    # with open('override_payload.json', 'w') as dest:
    #     json.dump(domain_data, dest)
    try:
        check_token_status()
        api_call_counter()
        resp = requests.post(
            f"https://{cfg.fmc_ip}/api/fmc_config/v1/domain/{domain_uuid}/object/securityzones?bulk=true",
            headers=cfg.headers_json, data=json.dumps(sz_payload), verify=False)
        # time.sleep(0.7)
        logging.info(f'request_body:\n{resp.request.body}')
        # logging.info(f'Headers: {str(resp.headers)}\n')
        logging.info(f'Text: {str(resp.text)}\n')
        logging.info(
            f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} security zones objects add operation status Code: {str(resp.status_code)}\n')
        if not str(resp.status_code).startswith('2'):
            errors_filename = 'outputs/errors.txt'
            with open(errors_filename, "a") as f:
                f.write(
                    f'Status code: {resp.status_code}\n Request.body: {resp.request.body}\n Request text: {resp.text}\n')
    except requests.exceptions.HTTPError as errh:
        logging.info(f'{errh}')
        raise SystemExit(errh)
    except requests.exceptions.RequestException as err:
        logging.info(f'{err}')
        raise SystemExit(err)
    

def post_portobjects_objects(portobjects_payload, domain_uuid):
    # with open('override_payload.json', 'w') as dest:
    #     json.dump(domain_data, dest)
    try:
        check_token_status()
        api_call_counter()
        resp = requests.post(
            f"https://{cfg.fmc_ip}/api/fmc_config/v1/domain/{domain_uuid}/object/protocolportobjects?bulk=true",
            headers=cfg.headers_json, data=json.dumps(portobjects_payload), verify=False)
        # time.sleep(0.7)
        logging.info(f'request_body:\n{resp.request.body}')
        # logging.info(f'Headers: {str(resp.headers)}\n')
        logging.info(f'Text: {str(resp.text)}\n')
        logging.info(
            f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} protocolportobjects add operation status Code: {str(resp.status_code)}\n')
        if not str(resp.status_code).startswith('2'):
            errors_filename = 'outputs/errors.txt'
            with open(errors_filename, "a") as f:
                f.write(
                    f'Status code: {resp.status_code}\n Request.body: {resp.request.body}\n Request text: {resp.text}\n')
    except requests.exceptions.HTTPError as errh:
        logging.info(f'{errh}')
        raise SystemExit(errh)
    except requests.exceptions.RequestException as err:
        logging.info(f'{err}')
        raise SystemExit(err)


def security_zones_operation(ws):
    """
    security_zones_operation create, delete, modify security zones method

    :param ws: xlsx sheet name
    :type ws: class openpyxl
    """

    max_row = ws.max_row
    max_column = ws.max_column
    
    add_zones_payload = list()
    add_domain_zones = dict()
    domain_zones_list = list()

    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to check/add/remove/modify Security Zones\n')
    
    for i in range(2, max_row+1):
        
        source_dest_zones_exist = all([
            ws.cell(row=i, column=3).value, 
            ws.cell(row=i, column=4).value, 
            ws.cell(row=i, column=5).value,
            ws.cell(row=i, column=10).value == 'add'])
        
        if source_dest_zones_exist:
            zone_domain = (ws.cell(row=i, column=3).value).strip()
            source_zone = (ws.cell(row=i, column=4).value).strip()
            dest_zone = (ws.cell(row=i, column=5).value).strip()
            zone_domain = get_domain_from_sz(zone_domain)
            if not source_zone in list(all_sz):
                add_zones_payload.append((zone_domain, source_zone))
            if not dest_zone in list(all_sz):
                add_zones_payload.append((zone_domain, dest_zone))
    if add_zones_payload:
        add_zones_payload = list(set(add_zones_payload))
        # for zone_domain, zone in add_zones_payload:
        #     domain_zones_list.append((zone_domain, zone))
        domain_zones_dict = sort_zones_by_domain(add_zones_payload)
        domain_sz_json = create_json_add_sz(domain_zones_dict)
        for domain, sz_payload in domain_sz_json.items():
            domain_uuid = get_domain_uuid(domain)['uuid']
            post_sz_objects(sz_payload, domain_uuid)
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished to check/add/remove/modify Security Zones\n')


def check_portobject_spelling(portobject):
    correct_protocols = ['tcp', 'udp', 'icmp']
    if not portobject in correct_protocols:
        for protocol in correct_protocols:
            match = 0
            for letter in portobject:
                if match == 2:
                    portobject = protocol
                    return portobject
                if letter in protocol and match != 2:
                    match += 1
                    continue
    else:
        return portobject
    

def parsed_portobjects(raw_portobject):
    portobjects = raw_portobject.split(',')
    prepared_portobjects = list()
    for portobject in portobjects:
        portobject = portobject.strip()
        portobject = portobject.replace('/', '.')
        match_two_proto = re.search(
            r'(?P<first>.+?(?=\-))'
            r'(?:\-)'
            r'(?P<second>.+?(?=\.))'
            r'(?:\.)'
            r'(?P<port>\d+)', portobject)
        match_two_proto_extended = re.search(
            r'(?P<first>.+?(?=\-))'
            r'(?:\-)'
            r'(?P<second>.+?(?=\.))'
            r'(?:\.)'
            r'(?P<ports>\d+\-\d+)', portobject)
        match_ports_range = re.search(
            r'(?P<protocol>.+?(?=\.))'
            r'(?:\.)'
            r'(?P<ports>\d+\-\d+)', portobject)
        match_protocol_port = re.search(
            r'(?P<protocol>.+?(?=\.))'
            r'(?:\.)'
            r'(?P<port>\d+)', portobject)
        if match_two_proto:
            first = check_portobject_spelling(match_two_proto.group("first"))
            second = check_portobject_spelling(match_two_proto.group("second"))
            prepared_portobjects.append(f'{first}.{match_two_proto.group("port")}')
            prepared_portobjects.append(f'{second}.{match_two_proto.group("port")}')
        elif match_two_proto_extended:
            first = check_portobject_spelling(match_two_proto_extended.group("first"))
            second = check_portobject_spelling(match_two_proto_extended.group("second"))
            prepared_portobjects.append(
                f'{first}.{match_two_proto_extended.group("ports")}')
            prepared_portobjects.append(
                f'{second}.{match_two_proto_extended.group("ports")}')
        elif match_ports_range:
            protocol = check_portobject_spelling(
                match_ports_range.group("protocol"))
            prepared_portobjects.append(
                f'{protocol}.{match_ports_range.group("ports")}')
        elif match_protocol_port:
            protocol = check_portobject_spelling(
                match_protocol_port.group("protocol"))
            prepared_portobjects.append(
                f'{protocol}.{match_protocol_port.group("port")}')
        elif portobject.lower() == 'icmp':
            prepared_portobjects.append(portobject.lower())
        elif portobject.lower() == 'ip':    
            prepared_portobjects.append('any')
        else:
            prepared_portobjects.append(portobject.lower())
    return prepared_portobjects
    

def portobjects_operation(ws):
    """
    portobjects_operation create, delete, modify security portobjects method

    :param ws: xlsx sheet name
    :type ws: class openpyxl
    """

    max_row = ws.max_row
    max_column = ws.max_column

    add_portobjects_payload = list()

    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to check/add/remove/modify PortObjects\n')

    for i in range(2, max_row+1):
        domain_port_action_exist = all([
            ws.cell(row=i, column=3).value, 
            ws.cell(row=i, column=8).value, 
            ws.cell(row=i, column=10).value == 'add'])
        if domain_port_action_exist:
            portobject_domain = ws.cell(row=i, column=3).value
            raw_portobject = ws.cell(row=i, column=8).value
            portobjects = parsed_portobjects(raw_portobject)
            portobject_domain = get_domain_for_portobject(portobject_domain)
            for portobject in portobjects:
                if portobject.lower() == 'any':
                    continue
                elif not portobject in list(all_portobjects):
                    add_portobjects_payload.append((portobject_domain, portobject))
    if add_portobjects_payload:
        add_portobjects_payload = list(set(add_portobjects_payload))
        portobject_domain_dict = sort_portobjects_by_domain(add_portobjects_payload)
        domain_portobjects_json = create_json_add_portobjects(portobject_domain_dict)
        for domain, portobjects_payload in domain_portobjects_json.items():
            domain_uuid = cfg.global_domain_uuid
            post_portobjects_objects(portobjects_payload, domain_uuid)

    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished to check/add/remove/modify Security portobjects\n')


def acl_rules_operation(ws):
    """
    acl_rules_operation operation with acl rules

    :param ws: xlsx sheet name
    :type ws: class openpyxl
    """

    max_row = ws.max_row
    max_column = ws.max_column

    acl_rule_payload = list()

    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to check/add/remove/modify acl rules\n')

    for i in range(2, max_row+1):
        alc_rule_data_action_exist = all([
            ws.cell(row=i, column=3).value,
            ws.cell(row=i, column=4).value,
            ws.cell(row=i, column=5).value,
            ws.cell(row=i, column=6).value,
            ws.cell(row=i, column=7).value,
            ws.cell(row=i, column=8).value,
            ws.cell(row=i, column=10).value == 'add'])
        if alc_rule_data_action_exist:
            domain_name = ws.cell(row=i, column=3).value
            source_zone = ws.cell(row=i, column=4).value
            dest_zone = ws.cell(row=i, column=5).value
            raw_source_networks = ws.cell(row=i, column=6).value
            raw_dest_networks = ws.cell(row=i, column=7).value
            raw_portobjects = ws.cell(row=i, column=8).value
            portobjects = parsed_portobjects(raw_portobjects)
            portobject_domain = get_domain_for_portobject(portobjects)
            for portobject in portobjects:
                if portobject.lower() == 'ip':
                    continue
                elif not portobject in list(all_portobjects):
                    acl_rule_payload.append(
                        (portobject_domain, portobject))
    if acl_rule_payload:
        acl_rule_payload = list(set(acl_rule_payload))
        portobject_domain_dict = sort_portobjects_by_domain(acl_rule_payload)
        domain_portobjects_json = create_json_add_portobjects(portobject_domain_dict)
        for domain, portobjects_payload in domain_portobjects_json.items():
            domain_uuid = cfg.global_domain_uuid
            post_portobjects_objects(portobjects_payload, domain_uuid)

    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished to check/add/remove/modify Security portobjects\n')


def del_portobjects():
    """
    object_del function to delete all ports

    """
    ''' now to DEL our list of network ports '''
    logging.info(f"\nStarting to work to DELETE all protocol port objects\n")

    for domain, portobjects in all_portobjects_domain.items():
        for portobject, portobject_data in portobjects.items():
            if portobject in system_ports:
                continue
            else:
                port_id = portobject_data['id']
                domain_uuid = get_domain_uuid(domain)['uuid']
                try:
                    check_token_status()
                    api_call_counter()
                    resp = requests.delete(
                        f"https://{cfg.fmc_ip}/api/fmc_config/v1/domain/{domain_uuid}/object/protocolportobjects/{port_id}",
                        headers=cfg.headers_json, verify=False)
                    # logging.info(f'request_body:\n{resp.request.body}')
                    # logging.info(f'Headers: {str(resp.headers)}\n')
                    # logging.info(f'Text: {str(resp.text)}\n')
                    logging.info(
                        f'Protocolportobject {portobject} del operation status Code: {str(resp.status_code)}\n')
                    if not str(resp.status_code).startswith('2'):
                        errors_filename = 'outputs/errors.txt'
                        with open(errors_filename, "a") as f: 
                            f.write(
                                f'Status code: {resp.status_code}\n Request.body: {resp.request.body}\n Request text: {resp.text}\n')

                except requests.exceptions.HTTPError as errh:
                    logging.info(f'{errh}')
                    raise SystemExit(errh)
                except requests.exceptions.RequestException as err:
                    logging.info(f'{err}')
                    raise SystemExit(err)
            logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}Finished deleting all portobjects\n')


def create_xlsx_and_sheets(output_xlsx):
    """
    create_xlsx_and_sheets Create template Excel sheets with all configured domains

    :param output_xlsx: output Excel filename
    :type output_xlsx: xlsx object
    """
    domains_list = list()
    for domain_data in cfg.all_domains_json:
        domain_name = domain_data['name']
        domain_name = domain_name.replace('/', '.')
        domains_list.append(domain_name)

    sorted_domains_list = sorted(domains_list, key=len)

    wb = Workbook()
    wb.active
    for domain in sorted_domains_list:
        wb.create_sheet(f'{domain}')
        # ws = wb.create_sheet(f'{domain}.group.obj')
    wb.remove(wb.active)
    wb.save(output_xlsx)
    wb.close()


def get_sheet_name_for_domain(domain_name, all_sheets):
    """
    get_sheet_name_for_domain Transform Domain name for Excel

    :param domain_name: FMC domain name
    :type domain_name: str()
    :param all_sheets: list of str()
    :type all_sheets: list of domain names transformed for Excel
    :return: sheet name (relevant to requested domain)
    :rtype: str()
    """
    domain_for_sheet = domain_name.replace('/', '.')
    for sheet in all_sheets:
        if sheet == domain_for_sheet:
            return sheet


def get_all_acp():
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to get info of all ACP\n')

    acp_data = list()
    all_acp_domain = dict()
    all_acp = dict()
    
    offset_value = 0
    limit_value = 1000
    for domain in cfg.all_domains_json:
        temp_domain_acp = dict()
        try:
            check_token_status()
            api_call_counter()
            resp = requests.get(
                f"https://{cfg.fmc_ip}/api/fmc_config/v1/domain/{domain.get('uuid')}/policy/accesspolicies?expanded=true&offset={offset_value}&limit={limit_value}",
                headers=cfg.headers_json,
                verify=False)
            # time.sleep(0.7)
            logging.info(f'GET accesspolicies request for domain {domain.get("name")} status code {resp.status_code}')
            # time.sleep(0.7)
            # if str(resp.status_code).startswith('2'):
            #     try:
            #         acp_data += (json.loads(resp.text)['items'])
            #     except KeyError:
            #         logging.info(
            #             f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} No ACP received.\n Maybe none ACPs configured?')
            if str(resp.status_code).startswith('2'):
                try:
                    acp_data += (json.loads(resp.text)['items'])
                    try:
                        next_url = resp.json()['paging']['next'][0]
                    except KeyError as error:
                        logging.info(
                            f' {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} end of page')
                        next_url = ''
                    while next_url:
                        check_token_status()
                        api_call_counter()
                        resp = requests.get(next_url,
                                            headers=cfg.headers_json,
                                            verify=False)
                        acp_data += (json.loads(resp.text)['items'])
                        try:
                            next_url = resp.json()['paging']['next'][0]
                        except KeyError as error:
                            logging.info(
                                f' {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} end of page')
                            next_url = ''
                except KeyError:
                    logging.info(
                        f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} No ACP received.\n Maybe none ACPs configured?')

            elif not str(resp.status_code).startswith('2'):
                errors_filename = 'outputs/errors.txt'
                with open(errors_filename, "a") as f:
                    f.write(
                        f'Status code: {resp.status_code}\n Request.body: {resp.request.body}\n Request text: {resp.text}\n')
                    
        except requests.exceptions.HTTPError as errh:
            logging.info(f'{errh}')
        except requests.exceptions.RequestException as err:
            logging.info(f'{err}')

        for acp in acp_data:
            if acp['name'] in list(all_acp):
                continue
            else:
                all_acp.update({acp.get('name'): acp})
                temp_domain_acp.update({acp.get('name'): acp})
                
        all_acp_domain.update({domain.get('name'): temp_domain_acp})
                
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished collecting info of all ACP\n')
    with open('outputs/all_acp.yaml', 'w') as dest, open('outputs/all_acp_domain.yaml', 'w') as dest2:
        yaml.SafeDumper.ignore_aliases = lambda *args: True
        yaml.safe_dump(all_acp, dest, default_flow_style=False)
        yaml.safe_dump(all_acp_domain, dest2, default_flow_style=False)
    return all_acp, all_acp_domain


def get_all_acp_rules():
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to work to get info of all ACP rules\n')
    acl_rules_acp_domain = dict()
    
    offset_value = 0
    limit_value = 1000
    for domain, acps in all_acp_domain.items():
        temp_domain_acl = dict()
        for acp, acp_data in acps.items():
            acp_id = acp_data.get('id')
            acp_name = acp_data.get('name')
            # logging.info(f'acp = {acp_name}')
            try:
                check_token_status()
                api_call_counter()
                resp = requests.get(
                    f"https://{cfg.fmc_ip}/api/fmc_config/v1/domain/{get_domain_uuid(domain)['uuid']}/policy/accesspolicies/{acp_id}/accessrules?expanded=true&offset={offset_value}&limit={limit_value}",
                    headers=cfg.headers_json,
                    verify=False)
                # time.sleep(0.7)
                logging.info(
                    f"GET accesspolicies request for domain {get_domain_uuid(domain)['name']} status code {resp.status_code}")
                if str(resp.status_code).startswith('2'):
                    acp_rules_data = dict()
                    try:
                        acp_rules_data = json.loads(resp.text)['items']
                    except KeyError:
                        logging.info(
                            f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")} No ACP received.\n Maybe none ACPs configured?')

                    for rule in acp_rules_data:
                        if rule['metadata']['accessPolicy']['name'] == acp_name:
                            temp_domain_acl.setdefault(
                                acp_name, []).append(rule)
                    # logging.info(f'acp dict = {temp_domain_acl}')
                elif not str(resp.status_code).startswith('2'):
                    errors_filename = 'outputs/errors.txt'
                    with open(errors_filename, "a") as f:
                        f.write(
                            f'Status code: {resp.status_code}\n Request.body: {resp.request.body}\n Request text: {resp.text}\n')

            except requests.exceptions.HTTPError as errh:
                logging.info(f'{errh}')
            except requests.exceptions.RequestException as err:
                logging.info(f'{err}')          
        
        acl_rules_acp_domain.update({domain: temp_domain_acl})
            
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished collecting info of all ACP rules\n')

    with open('outputs/all_acl_rules_acp_domain.yaml', 'w') as dest:
        yaml.SafeDumper.ignore_aliases = lambda *args: True
        yaml.safe_dump(acl_rules_acp_domain, dest, default_flow_style=False)
    return acl_rules_acp_domain


def getLasKeyValue(d):
    for k, v in d.items():
        if v:
            yield from getLasKeyValue(v)
        else:
            yield k


def SetLastListToDict(dictionary):
    for key, value in dictionary.items():
        tmp_dict = dict()
        if isinstance(value, list):
            for item in value:
                tmp_dict.setdefault(item)
            dictionary.update({key: tmp_dict})
    dic_aux = [val_aux for val_aux in dictionary.values()
               if isinstance(val_aux, dict)]
    for i in dic_aux:
        SetLastListToDict(i)


def MergeLastDict(tmp_dict, old_dictionary):
    search_list = list(tmp_dict)
    for item in search_list:
        if item in old_dictionary.keys():
            old_dictionary.update({item: tmp_dict[item]})
    dic_aux = [val_aux for val_aux in old_dictionary.values()
               if isinstance(val_aux, dict)]
    for i in dic_aux:
        MergeLastDict(tmp_dict, i)


def BuildInheritanceTree(root_acp, acp_domain):
    SetLastListToDict(root_acp)
    roots_list = getLasKeyValue(root_acp)
    domain = list(root_acp)[0]
    for root_key in roots_list:
        try:
            del acp_domain[domain][root_key]
        except KeyError:
            pass
    roots_list = getLasKeyValue(root_acp)

    for acp_root_key in roots_list:
        tmp_dict = dict()
        for value in acp_domain[domain].values():
            if value['metadata']['parentPolicy']['name'] == acp_root_key:
                tmp_dict.setdefault(acp_root_key, []).append(value['name'])
        MergeLastDict(tmp_dict, root_acp)
    if acp_domain.get(domain):
        BuildInheritanceTree(root_acp, acp_domain)
    return root_acp


def inherit_tree_depth(it, count=0):
    """Depth of a nested dict.
    # Arguments
        it: a nested dict or list.
        count: a constant value used in internal calculations.
    # Returns
        Numeric value.
    """
    if isinstance(it, list):
        if any(isinstance(v, list) or isinstance(v, dict) for v in it):
            for v in it:
                if isinstance(v, list) or isinstance(v, dict):
                    return inherit_tree_depth(v, count + 1)
        else:
            return count
    elif isinstance(it, dict):
        if any(isinstance(v, list) or isinstance(v, dict) for v in it.values()):
            for v in it.values():
                if isinstance(v, list) or isinstance(v, dict):
                    return inherit_tree_depth(v, count + 1)
        else:
            return count
    else:
        return count


def ACP_inheritance():
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting to build ACP inheritance\n')
    
    acp_domain_copy = {dom: {k: v for k, v in acp.items()}
                       for dom, acp in all_acp_domain.items()}
    inherit_tree = dict()
    for domain, acp_info in acp_domain_copy.items():
        root_acp = dict()
        for key, value in acp_info.items():
            if value.get('metadata').get('inherit') == False:
                root_acp.setdefault(domain, []).append(key)
        if root_acp:
            root_acp = BuildInheritanceTree(root_acp, acp_domain_copy)
        else:
            errors_filename = 'outputs/errors.txt'
            with open(errors_filename, "a") as f:
                f.write(
                    f'Failed to build ACP inheritance tree for domain {domain}\n')
        if root_acp.get(domain):
            inherit_tree.update({domain: root_acp.get(domain)})

    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished collection of ACP inheritance\n') 
    return inherit_tree


def reversedEnumerate(l):
    return zip(range(len(l)-1, -1, -1), l)


def dict_generator(indict, pre=None):
    pre = pre[:] if pre else []
    if isinstance(indict, dict):
        for key, value in indict.items():
            if isinstance(value, dict):
                for d in dict_generator(value, pre + [key]):
                    yield d
            elif isinstance(value, list) or isinstance(value, tuple):
                for v in value:
                    for d in dict_generator(v, pre + [key]):
                        yield d
            else:
                yield pre + [key, value]
    else:
        yield pre + [indict]
        
        
def ACP_rules_to_xlsx(output_xlsx, inherit_tree):
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Starting writing info into Excel sheets\n')
    
    wb = Workbook()
    wb = load_workbook(output_xlsx, read_only=False)
    all_sheets = wb.sheetnames
    
    cell_align = Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
       
    for domain, acp_rules in acl_rules_acp_domain.items():
        # if not acl_rules_acp_domain.get(domain):
        #     sheet_for_deletion = wb[get_sheet_name_for_domain(domain, all_sheets)]
        #     wb.remove(sheet_for_deletion)
        #     continue
            
        ws = wb[get_sheet_name_for_domain(domain, all_sheets)]
                
        index_map2 = dict()
        list_for_index = list()
        
        """
        Inherit tree has to provide nested lists with ACP inheritance 
        """

        value = list(dict_generator(inherit_tree.get(domain)))
        if value:
            list_for_index = [elem for item in value for elem in item if isinstance(item, list) if elem]
            
            for indx, item in reversedEnumerate(list_for_index):
                index_map2[item] = indx
            
            acp_rules_list = list(acp_rules)
            sorted_acp = sorted(acp_rules_list, key=lambda pair: index_map2[pair])
            sorted_acp_rules = dict()
            for acp in sorted_acp:
                sorted_acp_rules.update({acp: acp_rules[acp]})

        index = 1
                
        ws.cell(row=index, column=1, value='ACP name').font = Font(b=True)
        ws.cell(row=index, column=2, value='rule_name').font = Font(b=True)
        ws.cell(row=index, column=3, value='ruleIndex').font = Font(b=True)
        ws.cell(row=index, column=4, value='enabled').font = Font(b=True)
        ws.cell(row=index, column=5, value='action').font = Font(b=True)
        ws.cell(row=index, column=6, value='sourceZones').font = Font(b=True)
        ws.cell(row=index, column=7, value='destinationZones').font = Font(b=True)
        ws.cell(row=index, column=8, value='sourceNetworks').font = Font(b=True)
        ws.cell(row=index, column=9, value='destinationNetworks').font = Font(b=True)
        ws.cell(row=index, column=10, value='sourcePorts').font = Font(b=True)
        ws.cell(row=index, column=11, value='destinationPorts').font = Font(b=True)
        ws.cell(row=index, column=12, value='urls').font = Font(b=True)
        ws.cell(row=index, column=13, value='ipsPolicy').font = Font(b=True)
        ws.cell(row=index, column=14, value='filePolicy').font = Font(b=True)
        ws.cell(row=index, column=15, value='sendEventsToFMC').font = Font(b=True)
        ws.cell(row=index, column=16, value='logBegin').font = Font(b=True)
        ws.cell(row=index, column=17, value='logEnd').font = Font(b=True)
        ws.cell(row=index, column=18, value='enableSyslog').font = Font(b=True)
        ws.cell(row=index, column=19, value='comment').font = Font(b=True)
        ws.cell(row=index, column=20, value='newComments').font = Font(b=True)

        for column in ascii_uppercase:
            if column == 'A':
                ws.column_dimensions[column].width = 25
            elif column == 'B':
                ws.column_dimensions[column].width = 40
            elif column == 'C':
                ws.column_dimensions[column].width = 5
            elif column == 'D':
                ws.column_dimensions[column].width = 9
            elif column == 'E':
                ws.column_dimensions[column].width = 8
            elif column == 'F':
                ws.column_dimensions[column].width = 30
            elif column == 'G':
                ws.column_dimensions[column].width = 30
            elif column == 'H':
                ws.column_dimensions[column].width = 65
            elif column == 'I':
                ws.column_dimensions[column].width = 65
            elif column == 'J':
                ws.column_dimensions[column].width = 11
            elif column == 'K':
                ws.column_dimensions[column].width = 25
            elif column == 'L':
                ws.column_dimensions[column].width = 8
            elif column == 'M':
                ws.column_dimensions[column].width = 8
            elif column == 'N':
                ws.column_dimensions[column].width = 8
            elif column == 'O':
                ws.column_dimensions[column].width = 8
            elif column == 'P':
                ws.column_dimensions[column].width = 8
            elif column == 'Q':
                ws.column_dimensions[column].width = 8
            elif column == 'R':
                ws.column_dimensions[column].width = 8
            elif column == 'S':
                ws.column_dimensions[column].width = 8
            elif column == 'T':
                ws.column_dimensions[column].width = 8
                
        if value:
            acp_rules = sorted_acp_rules
        elif not value:
            pass
        
        for acp_rule, acp_rule_data_list in acp_rules.items():
            for acp_rule_data in acp_rule_data_list:
                ws.cell(row=index + 1, column=1, value=acp_rule).alignment = cell_align
                ws.cell(row=index + 1, column=1).border = thin_border
                try:
                    ws.cell(row=index + 1, column=2,value=acp_rule_data.get('name')).alignment = cell_align
                    ws.cell(row=index + 1, column=2).border = thin_border
                except (KeyError, TypeError):
                    logging.info(f'no rule name received')
                try:    
                    ws.cell(row=index + 1, column=3, value=acp_rule_data.get('metadata').get('ruleIndex')).alignment = cell_align
                    ws.cell(row=index + 1, column=3).border = thin_border
                except (KeyError, TypeError):
                    logging.info(f'no rule index received')
                try:
                    ws.cell(row=index + 1, column=4, value=acp_rule_data.get('enabled')).alignment = cell_align
                    ws.cell(row=index + 1, column=4).border = thin_border
                except (KeyError, TypeError):
                    logging.info(f'no rule enabled/disabled received')
                try:
                    ws.cell(row=index + 1, column=5, value=acp_rule_data.get('action')).alignment = cell_align
                    ws.cell(row=index + 1, column=5).border = thin_border
                except (KeyError, TypeError):
                    logging.info(f'no rule action received')
                try:
                    sz_json = acp_rule_data['sourceZones']['objects']
                    sz_list = [sz.get('name') for sz in sz_json]
                    sz_str = ', '.join(sz_list)
                    ws.cell(row=index + 1, column=6, value=sz_str).alignment = cell_align
                    ws.cell(row=index + 1, column=6).border = thin_border
                except (KeyError, TypeError):
                    logging.info(f'source SZ were not present. Assuming it was selected as any')
                    ws.cell(row=index + 1, column=6, value='any').alignment = cell_align
                    ws.cell(row=index + 1, column=6).border = thin_border
                try:
                    dz_json = acp_rule_data['destinationZones']['objects']
                    dz_list = [dz.get('name') for dz in dz_json]
                    dz_str = ', '.join(dz_list)
                    ws.cell(row=index + 1, column=7, value=dz_str).alignment = cell_align
                    ws.cell(row=index + 1, column=7).border = thin_border
                except (KeyError, TypeError):
                    logging.info(
                        f'destination SZ were not present. Assuming it was selected as any')
                    ws.cell(row=index + 1, column=7, value='any').alignment = cell_align
                    ws.cell(row=index + 1, column=7).border = thin_border
                try:
                    source_net_json = acp_rule_data['sourceNetworks']['objects']
                    source_net_list = [source_net.get('name') for source_net in source_net_json]
                    source_net_str = ', '.join(source_net_list)
                    ws.cell(row=index + 1, column=8, value=source_net_str).alignment = cell_align
                    ws.cell(row=index + 1, column=8).border = thin_border
                except (KeyError, TypeError):
                    logging.info(
                        f'source Network were not present. Assuming it was selected as any')
                    ws.cell(row=index + 1, column=8, value='any').alignment = cell_align
                    ws.cell(row=index + 1, column=8).border = thin_border
                try:
                    dest_net_json = acp_rule_data['destinationNetworks']['objects']
                    dest_net_list = [dest_net.get('name') for dest_net in dest_net_json]
                    dest_net_str = ', '.join(dest_net_list)
                    ws.cell(row=index + 1, column=9, value=dest_net_str).alignment = cell_align
                    ws.cell(row=index + 1, column=9).border = thin_border
                except (KeyError, TypeError):
                    logging.info(
                        f'destination Network were not present. Assuming it was selected as any')
                    ws.cell(row=index + 1, column=9, value='any').alignment = cell_align
                    ws.cell(row=index + 1, column=9).border = thin_border
                try:
                    source_port_json = acp_rule_data['sourcePorts']['objects']
                    source_port_list = [source_port.get('name') for source_port in source_port_json]
                    source_port_str = ', '.join(source_port_list)
                    ws.cell(row=index + 1, column=10, value=source_port_str).alignment = cell_align
                    ws.cell(row=index + 1, column=10).border = thin_border
                except (KeyError, TypeError):
                    logging.info(
                        f'source PortObject were not present. Assuming it was selected as any')
                    ws.cell(row=index + 1, column=10, value='any').alignment = cell_align
                    ws.cell(row=index + 1, column=10).border = thin_border
                if acp_rule_data.get('destinationPorts'):
                    if acp_rule_data.get('destinationPorts').get('objects'):
                        try:
                            dest_port_json = acp_rule_data['destinationPorts']['objects']
                            dest_port_list = [dest_port.get('name') for dest_port in dest_port_json]
                            dest_port_str = ', '.join(dest_port_list)
                            ws.cell(row=index + 1, column=11,value=dest_port_str).alignment = cell_align
                            ws.cell(row=index + 1, column=11).border = thin_border
                        except (KeyError, TypeError):
                            logging.info(f'destination PortObject were not present. Assuming it was selected as any')
                            ws.cell(row=index + 1, column=11, value='any').alignment = cell_align
                            ws.cell(row=index + 1, column=11).border = thin_border
                    elif acp_rule_data.get('destinationPorts').get('literals'):
                        try:
                            dest_port_json = acp_rule_data['destinationPorts']['literals']
                            # dest_port_list = [zip(system_literals_map.get(dest_port.get('protocol')), dest_port.get('port')) for dest_port in dest_port_json]
                            # dest_port_list = [(system_literals_map.get(dest_port.get('protocol')) + '.' + dest_port.get('port')) for dest_port in dest_port_json]
                            dest_port_list = [(system_literals_map.get(dest_port.get('protocol')) + '.' + (dest_port.get('icmpType') if 'icmpType' in dest_port else dest_port.get('port'))) for dest_port in dest_port_json]
                            dest_port_str = ', '.join(dest_port_list)
                            ws.cell(row=index + 1, column=11,value=dest_port_str).alignment = cell_align
                            ws.cell(row=index + 1, column=11).border = thin_border
                        except (KeyError, TypeError):
                            logging.info(f'destination PortObject were not present. Assuming it was selected as any')
                            ws.cell(row=index + 1, column=11, value='any').alignment = cell_align
                            ws.cell(row=index + 1, column=11).border = thin_border
                else:
                    logging.info(f'destination PortObject were not present. Assuming it was selected as any')
                    ws.cell(row=index + 1, column=11, value='any').alignment = cell_align
                    ws.cell(row=index + 1, column=11).border = thin_border
                try:
                    url_list = acp_rule_data.get('urls')['objects']
                    urls = [url.get('name') for url in url_list]
                    urls_str = ', '.join(urls)
                    ws.cell(row=index + 1, column=12, value=urls_str).alignment = cell_align
                    ws.cell(row=index + 1, column=12).border = thin_border
                except (KeyError, TypeError):
                    logging.info(
                        f'url were not present')
                    pass
                try:
                    ws.cell(row=index + 1, column=13, value=acp_rule_data.get('ipsPolicy').get('name')).alignment = cell_align
                    ws.cell(row=index + 1, column=13).border = thin_border
                except (KeyError, TypeError, AttributeError):
                    logging.info(
                        f'ipsPolicy were not present')
                    pass
                try:
                    ws.cell(row=index + 1, column=14, value=acp_rule_data.get('filePolicy').get('name')).alignment = cell_align
                    ws.cell(row=index + 1, column=14).border = thin_border
                except (KeyError, TypeError, AttributeError):
                    logging.info(
                        f'filePolicy were not present')
                    pass
                ws.cell(row=index + 1, column=15, value=acp_rule_data.get('sendEventsToFMC')).alignment = cell_align
                ws.cell(row=index + 1, column=15).border = thin_border
                ws.cell(row=index + 1, column=16, value=acp_rule_data.get('logBegin')).alignment = cell_align
                ws.cell(row=index + 1, column=16).border = thin_border
                ws.cell(row=index + 1, column=17, value=acp_rule_data.get('logEnd')).alignment = cell_align
                ws.cell(row=index + 1, column=17).border = thin_border
                ws.cell(row=index + 1, column=18, value=acp_rule_data.get('enableSyslog')).alignment = cell_align
                ws.cell(row=index + 1, column=18).border = thin_border
                try:
                    comments = acp_rule_data.get('commentHistoryList')
                    comments_list = [comment.get('comment') for comment in comments]
                    comment_str = ', '.join(comments_list)
                    ws.cell(row=index + 1, column=19, value=comment_str).alignment = cell_align
                    ws.cell(row=index + 1, column=19).border = thin_border
                except (KeyError, TypeError):
                    logging.info(
                        f'comments were not present')
                    pass
        
                index += 1        
        
        if inherit_tree_depth(inherit_tree) > 1:
            index += 3
            ws.cell(row=index, column=1,value='ACP inheritance').font = Font(b=True)
            index += 1
            
            inherit_list = list(dict_generator(inherit_tree.get(domain)))           
            if inherit_list:
                for acp_blocks in inherit_list:
                    j = 1
                    if isinstance(acp_blocks, list):
                        for acp_name in acp_blocks:
                            if isinstance(acp_name, str):
                                ws.cell(row=index, column=j,value=acp_name).alignment = cell_align
                                ws.cell(row=index, column=j,value=acp_name).border = thin_border
                                j += 1
                                index += 1

    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Finished writing info into Excel sheets\n')
    wb.save(output_xlsx) 
    wb.close()


if __name__ == "__main__":
    cfg.init()
    output_dir = 'outputs'
    try:
        os.makedirs(output_dir)
    except OSError as e:
        pass
    errors_filename = 'outputs/errors.txt'
    with open(errors_filename, "w") as f:
        f.write('')

    cfg.all_domains_json = get_all_domains_data()
    cfg.all_obj_domain, cfg.all_ids_domain = get_all_objects_with_domains()
    cfg.all_devices = get_all_devices()
    cfg.all_detailed_networkgroups = get_all_groups_info()
    system_ports = cfg.system_ports
    
    system_literals_map = dict()
    
    system_literals_map = {'6': 'tcp',
                           '17': 'udp',
                           '1': 'icmp'}

    all_sz, all_sz_domain = get_all_sz()
    all_portobjects, all_portobjects_domain = get_all_portobjects()
    all_acp, all_acp_domain = get_all_acp()
    acl_rules_acp_domain = get_all_acp_rules()
    
    ''' read and parse data out of the XLSX Commutation map '''

    create_xlsx_and_sheets(cfg.output_acp_xlsx)
    inherit_tree = ACP_inheritance()
    ACP_rules_to_xlsx(cfg.output_acp_xlsx, inherit_tree)
    logging.info(f'\n {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} Download of ACP,ACL has been completed with no major exceptions.\n File is {cfg.output_acp_xlsx} \n Done!\n')
